import sys
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import xml.etree.ElementTree as ET
import ast
from typing import Dict, List, Tuple, Optional

# comp_xml.py와 comp_q.py에서 필요한 클래스와 함수 import
from comp_xml import read_interface_block, XMLComparator
from comp_q import QueryParser

class InterfaceXMLToExcel:
    def __init__(self, excel_path: str, xml_dir: str, output_path: str = 'test24.xlsx'):
        """
        XML 파일에서 추출한 쿼리의 컬럼과 VALUES를 매핑하여 Excel 파일을 생성하는 클래스
        
        Args:
            excel_path (str): 인터페이스 정보가 있는 Excel 파일 경로
            xml_dir (str): XML 파일이 있는 디렉토리 경로
            output_path (str): 출력할 Excel 파일 경로
        """
        self.excel_path = excel_path
        self.xml_dir = xml_dir
        self.output_path = output_path
        self.query_parser = QueryParser()
        
        # Excel 파일 로드
        self.input_workbook = openpyxl.load_workbook(excel_path)
        self.input_worksheet = self.input_workbook.active
        
        # 출력 Excel 파일 생성
        self.output_workbook = openpyxl.Workbook()
        self.output_worksheet = self.output_workbook.active
        
        # 스타일 정의
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True, size=9)
        self.normal_font = Font(name='맑은 고딕', size=9)
        self.bold_font = Font(bold=True, size=9)
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    def find_rcv_file(self, if_id: str) -> str:
        """
        주어진 인터페이스 ID에 해당하는 수신(RCV) XML 파일을 찾습니다.
        
        Args:
            if_id (str): 인터페이스 ID
            
        Returns:
            str: 찾은 파일의 경로, 없으면 None
        """
        if not if_id:
            print(f"Warning: Empty IF_ID provided")
            return None
            
        try:
            # 디렉토리 내의 모든 XML 파일 검색
            for file in os.listdir(self.xml_dir):
                if not file.startswith(if_id):
                    continue
                    
                # 수신 파일 (.RCV.xml)
                if file.endswith('.RCV.xml'):
                    file_path = os.path.join(self.xml_dir, file)
                    return file_path
            
            print(f"Warning: No receive file found for IF_ID: {if_id}")
            return None
            
        except Exception as e:
            print(f"Error finding interface files: {e}")
            return None
    
    def extract_query_from_xml(self, xml_path: str) -> str:
        """
        XML 파일에서 SQL 쿼리를 추출합니다.
        
        Args:
            xml_path (str): XML 파일 경로
            
        Returns:
            str: 추출된 SQL 쿼리, 없으면 None
        """
        try:
            # XML 파일이 제대로 로드되었는지 확인
            if not os.path.exists(xml_path):
                print(f"Warning: XML file not found: {xml_path}")
                return None
                
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # XML 내용이 유효한지 확인
            if root is None:
                print(f"Warning: Invalid XML content in file: {xml_path}")
                return None
            
            # SQL 노드 찾기
            sql_node = root.find(".//SQL")
            if sql_node is None or not sql_node.text:
                print(f"Warning: No SQL content found in file: {xml_path}")
                return None
                
            query = sql_node.text.strip()
            
            # 추출된 쿼리가 유효한지 확인
            if not query:
                print(f"Warning: Empty SQL query in file: {xml_path}")
                return None
                
            return query
            
        except ET.ParseError as e:
            print(f"Error parsing XML file {xml_path}: {e}")
            return None
        except Exception as e:
            print(f"Unexpected error processing file {xml_path}: {e}")
            return None
    
    def clean_value(self, value: str) -> str:
        """
        VALUES 항목에서 콜론(:)을 제거하고 TO_DATE 함수 내의 실제 값만 추출합니다.
        
        Args:
            value (str): 원본 값
            
        Returns:
            str: 정제된 값
        """
        if not value:
            return ""
        
        # 콜론(:) 제거
        cleaned_value = value.replace(':', '')
        
        # TO_DATE 함수 처리
        to_date_pattern = r'TO_DATE\(\s*:?([A-Za-z0-9_]+)\s*,\s*[\'"](.*?)[\'"](\s*\))'
        to_date_match = re.search(to_date_pattern, value, re.IGNORECASE)
        
        if to_date_match:
            # TO_DATE 함수에서 첫 번째 인자(실제 값)만 추출
            param_name = to_date_match.group(1)
            return param_name
        
        return cleaned_value
    
    def get_column_value_mapping(self, query: str) -> Dict[str, str]:
        """
        INSERT 쿼리에서 컬럼과 VALUES를 매핑합니다.
        
        Args:
            query (str): INSERT SQL 쿼리
            
        Returns:
            Dict[str, str]: 컬럼과 값의 매핑 딕셔너리
        """
        if not query:
            return {}
            
        # QueryParser를 사용하여 INSERT 쿼리 파싱
        insert_parts = self.query_parser.parse_insert_parts(query)
        if not insert_parts:
            print(f"Failed to parse INSERT query: {query}")
            return {}
            
        # 테이블 이름과 컬럼-값 매핑 추출
        table_name, columns = insert_parts
        
        # 값 정제 처리
        cleaned_columns = {}
        for col, val in columns.items():
            cleaned_columns[col] = self.clean_value(val)
            
        return cleaned_columns
    
    def process_interfaces(self):
        """
        Excel 파일에서 인터페이스 정보를 읽고, XML 파일에서 쿼리를 추출하여 매핑 후 출력 Excel 파일에 작성합니다.
        """
        # 헤더 행 복사 및 스타일 적용
        for col in range(1, self.input_worksheet.max_column + 1):
            self.output_worksheet.cell(row=1, column=col).value = self.input_worksheet.cell(row=1, column=col).value
            self.output_worksheet.cell(row=1, column=col).font = self.normal_font
            
            self.output_worksheet.cell(row=2, column=col).value = self.input_worksheet.cell(row=2, column=col).value
            self.output_worksheet.cell(row=2, column=col).font = self.normal_font
            
            self.output_worksheet.cell(row=3, column=col).value = self.input_worksheet.cell(row=3, column=col).value
            self.output_worksheet.cell(row=3, column=col).font = self.normal_font
            
            self.output_worksheet.cell(row=4, column=col).value = self.input_worksheet.cell(row=4, column=col).value
            self.output_worksheet.cell(row=4, column=col).font = self.normal_font
        
        # 인터페이스 블록 처리
        current_col = 2  # B열부터 시작
        interface_count = 0
        
        while current_col <= self.input_worksheet.max_column:
            try:
                # 인터페이스 정보 읽기
                interface_info = read_interface_block(self.input_worksheet, current_col)
                if not interface_info:
                    break
                    
                interface_count += 1
                interface_id = interface_info.get('interface_id', '')
                interface_name = interface_info.get('interface_name', f'Interface_{interface_count}')
                
                print(f"\n처리 중인 인터페이스: {interface_name} (ID: {interface_id})")
                
                # 인터페이스 기본 정보 복사 및 스타일 적용
                self.output_worksheet.cell(row=1, column=current_col).value = interface_name
                self.output_worksheet.cell(row=1, column=current_col).font = self.normal_font
                
                self.output_worksheet.cell(row=2, column=current_col).value = interface_id
                self.output_worksheet.cell(row=2, column=current_col).font = self.normal_font
                
                self.output_worksheet.cell(row=3, column=current_col).value = self.input_worksheet.cell(row=3, column=current_col).value
                self.output_worksheet.cell(row=3, column=current_col).font = self.normal_font
                self.output_worksheet.cell(row=3, column=current_col + 1).value = self.input_worksheet.cell(row=3, column=current_col + 1).value
                self.output_worksheet.cell(row=3, column=current_col + 1).font = self.normal_font
                
                self.output_worksheet.cell(row=4, column=current_col).value = self.input_worksheet.cell(row=4, column=current_col).value
                self.output_worksheet.cell(row=4, column=current_col).font = self.normal_font
                self.output_worksheet.cell(row=4, column=current_col + 1).value = self.input_worksheet.cell(row=4, column=current_col + 1).value
                self.output_worksheet.cell(row=4, column=current_col + 1).font = self.normal_font
                
                # 수신 XML 파일 찾기
                rcv_file_path = self.find_rcv_file(interface_id)
                if not rcv_file_path:
                    print(f"Warning: No receive file found for interface {interface_name} (ID: {interface_id})")
                    current_col += 3  # 다음 인터페이스로 이동
                    continue
                
                # XML 파일에서 쿼리 추출
                query = self.extract_query_from_xml(rcv_file_path)
                if not query:
                    print(f"Warning: Failed to extract query from file {rcv_file_path}")
                    current_col += 3  # 다음 인터페이스로 이동
                    continue
                
                # 쿼리에서 컬럼-값 매핑 추출
                column_value_mapping = self.get_column_value_mapping(query)
                if not column_value_mapping:
                    print(f"Warning: Failed to extract column-value mapping from query")
                    current_col += 3  # 다음 인터페이스로 이동
                    continue
                
                # 특수 컬럼 제외
                special_columns = set(self.query_parser.special_columns['recv']['required'])
                filtered_mapping = {k: v for k, v in column_value_mapping.items() if k.upper() not in special_columns}
                
                # 매핑 정보를 Excel에 작성
                row = 5  # 5행부터 컬럼 매핑 시작
                for column, value in filtered_mapping.items():
                    # 수신 컬럼을 첫 번째 열(B열)에 배치
                    self.output_worksheet.cell(row=row, column=current_col).value = column  # 수신 컬럼을 첫 번째 열에 배치
                    self.output_worksheet.cell(row=row, column=current_col).font = self.normal_font
                    
                    # VALUES 항목을 오른쪽 열(C열)에 배치 - 이미 정제된 값 사용
                    self.output_worksheet.cell(row=row, column=current_col + 1).value = value  # VALUES 항목 (콜론 제거와 TO_DATE 함수 처리가 적용됨)
                    self.output_worksheet.cell(row=row, column=current_col + 1).font = self.normal_font
                    
                    row += 1
                
                print(f"인터페이스 {interface_name} (ID: {interface_id}) 처리 완료")
                
            except Exception as e:
                print(f"Error processing interface at column {current_col}: {str(e)}")
            
            current_col += 3  # 다음 인터페이스로 이동
        
        # 출력 파일 저장
        self.output_workbook.save(self.output_path)
        self.input_workbook.close()
        self.output_workbook.close()
        
        print(f"\n=== 처리 완료 ===")
        print(f"총 처리된 인터페이스 수: {interface_count}")
        print(f"출력 파일 저장 완료: {self.output_path}")

def main():
    try:
        # 하드코딩된 경로 설정 (python test24.py만 실행해도 작동하도록)
        # 현재 스크립트가 있는 디렉토리 기준으로 상대 경로 설정
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 기본 경로 설정 (실제 환경에 맞게 수정 필요)
        excel_path = os.path.join(current_dir, 'C:\\work\\LT\\input_W7.xlsx')  # 인터페이스 정보 파일
        xml_dir = os.path.join(current_dir, 'C:\\work\\LT\\W7xml')  # XML 파일 디렉토리
        output_path = os.path.join(current_dir, 'C:\\work\\LT\\test24.xlsx')  # 출력 파일
        
        # 명령행 인수가 있으면 덮어쓰기
        if len(sys.argv) > 1:
            excel_path = sys.argv[1]
        if len(sys.argv) > 2:
            xml_dir = sys.argv[2]
        if len(sys.argv) > 3:
            output_path = sys.argv[3]
            
        print(f"사용할 파일 경로:")
        print(f"- 인터페이스 정보 파일: {excel_path}")
        print(f"- XML 파일 디렉토리: {xml_dir}")
        print(f"- 출력 파일: {output_path}")
        
        # 인터페이스 처리 실행
        processor = InterfaceXMLToExcel(excel_path, xml_dir, output_path)
        processor.process_interfaces()
        
    except Exception as e:
        print(f"\n[심각한 오류] 프로그램 실행 중 오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()