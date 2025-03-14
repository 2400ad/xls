import openpyxl
import ast
from maptest import ColumnMapper
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def safe_get_dict_value(dictionary, key, default=''):
    """
    딕셔너리에서 안전하게 값을 가져오는 유틸리티 함수
    
    Args:
        dictionary: 검색할 딕셔너리 (None일 수 있음)
        key: 검색할 키
        default: 기본값 (dictionary가 None이거나 key가 없을 경우 반환)
    
    Returns:
        찾은 값 또는 기본값
    """
    if dictionary is None:
        return default
    return dictionary.get(key, default)

def read_interface_block(ws, start_col):
    """Excel에서 3컬럼 단위로 하나의 인터페이스 정보를 읽습니다."""
    interface_info = {
        'interface_name': ws.cell(row=1, column=start_col).value or '',  # IF NAME (1행)
        'interface_id': ws.cell(row=2, column=start_col).value or '',    # IF ID (2행)
        'send': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None},
        'recv': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None}
    }
    
    try:
        # DB 연결 정보 (3행에서 읽기)
        send_db_info = ast.literal_eval(ws.cell(row=3, column=start_col).value)
        recv_db_info = ast.literal_eval(ws.cell(row=3, column=start_col + 1).value)
        interface_info['send']['db_info'] = send_db_info
        interface_info['recv']['db_info'] = recv_db_info
        
        # 테이블 정보 (4행에서 읽기)
        send_table_info = ast.literal_eval(ws.cell(row=4, column=start_col).value)
        recv_table_info = ast.literal_eval(ws.cell(row=4, column=start_col + 1).value)
        
        interface_info['send']['owner'] = send_table_info.get('owner')
        interface_info['send']['table_name'] = send_table_info.get('table_name')
        interface_info['recv']['owner'] = recv_table_info.get('owner')
        interface_info['recv']['table_name'] = recv_table_info.get('table_name')
        
        # 컬럼 매핑 정보 (5행부터)
        row = 5
        while True:
            send_col = ws.cell(row=row, column=start_col).value
            recv_col = ws.cell(row=row, column=start_col + 1).value
            
            if not send_col and not recv_col:
                break
                
            interface_info['send']['columns'].append(send_col if send_col else '')
            interface_info['recv']['columns'].append(recv_col if recv_col else '')
            row += 1
            
    except Exception as e:
        print(f'인터페이스 정보 읽기 중 오류 발생: {str(e)}')
        return None
    
    return interface_info

def process_interface(interface_info, mapper):
    """하나의 인터페이스에 대한 모든 처리를 수행합니다."""
    if not interface_info:
        return {
            'comparison': None,
            'send_sql': None,
            'recv_sql': None,
            'field_xml': None,
            'errors': ['인터페이스 정보를 읽을 수 없습니다.']
        }
    
    results = {
        'comparison': None,
        'send_sql': None,
        'recv_sql': None,
        'field_xml': None,
        'errors': []
    }
    
    try:
        # DB 연결
        if interface_info['send']['db_info']:
            mapper.connect_send_db(
                interface_info['send']['db_info']['sid'],
                interface_info['send']['db_info']['username'],
                interface_info['send']['db_info']['password']
            )
        else:
            results['errors'].append("송신 DB 연결 정보가 없습니다.")
            return results
            
        if interface_info['recv']['db_info']:
            mapper.connect_recv_db(
                interface_info['recv']['db_info']['sid'],
                interface_info['recv']['db_info']['username'],
                interface_info['recv']['db_info']['password']
            )
        else:
            results['errors'].append("수신 DB 연결 정보가 없습니다.")
            return results

        # 테이블 정보 설정
        if interface_info['send']['owner'] and interface_info['send']['table_name']:
            mapper.set_send_table(interface_info['send']['owner'], interface_info['send']['table_name'])
        else:
            results['errors'].append("송신 테이블 정보가 없습니다.")
            return results
            
        if interface_info['recv']['owner'] and interface_info['recv']['table_name']:
            mapper.set_recv_table(interface_info['recv']['owner'], interface_info['recv']['table_name'])
        else:
            results['errors'].append("수신 테이블 정보가 없습니다.")
            return results
        
        # 매핑 설정
        send_columns = '\n'.join(interface_info['send']['columns'])
        recv_columns = '\n'.join(interface_info['recv']['columns'])
        mapper.set_send_mapping(send_columns)
        mapper.set_recv_mapping(recv_columns)
        
        # 컬럼 비교 수행
        results['comparison'] = mapper.compare_columns()
        
        # SQL 생성
        results['send_sql'] = mapper.generate_send_sql_from_mapping()
        results['recv_sql'] = mapper.generate_recv_sql()
        
        # 필드 XML 생성
        results['field_xml'] = mapper.generate_field_xml_from_mapping()
        
    except Exception as e:
        results['errors'].append(str(e))
    finally:
        mapper.close_connections()
    
    return results

def write_interface_result_to_sheet(wb, interface_info, results, interface_num):
    """각 인터페이스의 결과를 새로운 시트에 기록합니다."""
    interface_name = interface_info.get('interface_name', '').strip()
    interface_id = interface_info.get('interface_id', '').strip()
    sheet_name = f'{interface_num}_{interface_name}' if interface_name else f'Interface_{interface_num}'
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    # 스타일 정의
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=9)  # 헤더 폰트 크기 9
    normal_font = Font(name='맑은 고딕', size=9)  # 일반 텍스트 폰트 크기 9
    bold_font = Font(bold=True, size=9)  # 굵은 글씨 폰트 크기 9
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. 인터페이스 기본 정보 섹션
    ws['A1'] = '인터페이스 기본 정보'
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 25

    # 송신 테이블 정보
    ws['A2'] = '송신 테이블'
    ws['A2'].alignment = center_alignment
    ws.merge_cells('B2:E2')
    ws[f'B2'] = f"{interface_info['send']['owner']}.{interface_info['send']['table_name']}"
    ws[f'B2'].alignment = left_alignment

    # 수신 테이블 정보
    ws['A3'] = '수신 테이블'
    ws['A3'].alignment = center_alignment
    ws.merge_cells('B3:E3')
    ws[f'B3'] = f"{interface_info['recv']['owner']}.{interface_info['recv']['table_name']}"
    ws[f'B3'].alignment = left_alignment
    
    # 2. 컬럼 비교 결과 섹션
    ws['A4'] = '컬럼 비교 결과'
    ws['A4'].fill = header_fill
    ws['A4'].font = header_font
    ws.merge_cells('A4:J4')  # J열까지 확장
    ws['A4'].alignment = center_alignment
    ws.row_dimensions[4].height = 25
    
    comparison_headers = ['송신 컬럼', '송신 타입', '송신 크기', '송신 Null여부', 
                         '수신 컬럼', '수신 타입', '수신 크기', '수신 Null여부', 
                         '비교 결과', '상태']
    for idx, header in enumerate(comparison_headers):
        col = chr(ord('A') + idx)
        ws[f'{col}5'] = header
        ws[f'{col}5'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws[f'{col}5'].font = bold_font
        ws[f'{col}5'].alignment = center_alignment
        ws[f'{col}5'].border = border
    ws.row_dimensions[5].height = 20
    
    row = 6
    if results['comparison']:
        for comp in results['comparison']:
            send_info = comp.get('send_info', {})
            recv_info = comp.get('recv_info', {})
            
            # 송신 컬럼 정보
            ws[f'A{row}'] = comp.get('send_column', '')
            ws[f'B{row}'] = send_info.get('type', '')
            ws[f'C{row}'] = send_info.get('size', '')
            # 글자 타입이고 크기가 1024보다 큰 경우 노란색으로 표시
            if send_info.get('type', '').upper() in ['VARCHAR', 'VARCHAR2', 'CHAR']:
                try:
                    size = int(str(send_info.get('size', '0')).strip())
                    if size > 1024:
                        ws[f'C{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                except ValueError:
                    pass
            ws[f'D{row}'] = send_info.get('nullable', '')
            
            # 수신 컬럼 정보
            ws[f'E{row}'] = comp.get('recv_column', '')
            ws[f'F{row}'] = safe_get_dict_value(recv_info, 'type', '')
            ws[f'G{row}'] = safe_get_dict_value(recv_info, 'size', '')
            # 글자 타입이고 크기가 1024보다 큰 경우 노란색으로 표시
            if safe_get_dict_value(recv_info, 'type', '').upper() in ['VARCHAR', 'VARCHAR2', 'CHAR']:
                try:
                    size = int(str(safe_get_dict_value(recv_info, 'size', '0')).strip())
                    if size > 1024:
                        ws[f'G{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                except ValueError:
                    pass
            ws[f'H{row}'] = safe_get_dict_value(recv_info, 'nullable', '')
            
            # 비교 결과와 상태
            warnings = comp.get('warnings', [])
            errors = comp.get('errors', [])
            
            if errors:
                ws[f'I{row}'] = '\n'.join(errors)
                ws[f'J{row}'] = '확인필요'
                status_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            # elif warnings:
            #     ws[f'I{row}'] = ''  # 경고 메시지는 표시하지 않음
            #     ws[f'J{row}'] = '확인필요'
            #     status_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # 노란색
            else:
                ws[f'I{row}'] = ''
                ws[f'J{row}'] = '정상'
                status_fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')
            
            ws[f'J{row}'].fill = status_fill
            # 각 셀의 alignment 설정과 폰트 적용
            for col in range(ord('A'), ord('K')):
                col_letter = chr(col)
                cell = ws[f'{col_letter}{row}']
                cell.alignment = left_alignment if col_letter == 'I' else center_alignment
                cell.font = normal_font
                cell.border = border
            
            row += 1
    
    # 3. SQL 섹션
    current_row = row + 1
    ws['A{}'.format(current_row)] = 'SQL 문'
    ws['A{}'.format(current_row)].fill = header_fill
    ws['A{}'.format(current_row)].font = header_font
    ws.merge_cells('A{}:J{}'.format(current_row, current_row))
    ws['A{}'.format(current_row)].alignment = center_alignment
    ws.row_dimensions[current_row].height = 25

    # 송신/수신 SQL 헤더
    current_row += 1
    ws[f'A{current_row}'] = '송신 SQL'
    ws[f'F{current_row}'] = '수신 SQL'
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'F{current_row}'].alignment = center_alignment
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.merge_cells(f'F{current_row}:J{current_row}')

    # SQL 내용
    current_row += 1
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = results['send_sql']
    ws[f'A{current_row}'].font = normal_font
    ws[f'A{current_row}'].alignment = left_alignment

    ws.merge_cells(f'F{current_row}:J{current_row}')
    ws[f'F{current_row}'] = results['recv_sql']
    ws[f'F{current_row}'].font = normal_font
    ws[f'F{current_row}'].alignment = left_alignment

    # SQL 문의 줄 수에 따라 행 높이 조절 (2배로)
    lines_send = len(str(results['send_sql']).split('\n'))
    lines_recv = len(str(results['recv_sql']).split('\n'))
    max_lines = max(lines_send, lines_recv)
    ws.row_dimensions[current_row].height = max(25, min(30 * max_lines, 800))  # 기존 15 * max_lines에서 30 * max_lines로 변경

    # 4. XML 섹션
    current_row += 2
    ws['A{}'.format(current_row)] = '필드 XML'
    ws['A{}'.format(current_row)].fill = header_fill
    ws['A{}'.format(current_row)].font = header_font
    ws.merge_cells('A{}:J{}'.format(current_row, current_row))
    ws['A{}'.format(current_row)].alignment = center_alignment
    ws.row_dimensions[current_row].height = 25

    # XML 내용
    current_row += 1
    ws.merge_cells(f'A{current_row}:J{current_row}')
    if results.get('field_xml'):
        xml_lines = results['field_xml'].split('\n')
        if xml_lines and '<?xml' in xml_lines[0]:
            results['field_xml'] = '\n'.join(xml_lines[1:]).strip()
    ws[f'A{current_row}'] = results['field_xml']
    ws[f'A{current_row}'].font = normal_font
    ws[f'A{current_row}'].alignment = left_alignment

    # XML 문의 줄 수에 따라 행 높이 조절 (2배로)
    lines = len(str(results['field_xml']).split('\n'))
    ws.row_dimensions[current_row].height = max(25, min(30 * lines, 800))  # 기존 15 * lines에서 30 * lines로 변경
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20  # 송신 컬럼
    ws.column_dimensions['B'].width = 15  # 송신 타입
    ws.column_dimensions['C'].width = 12  # 송신 크기
    ws.column_dimensions['D'].width = 12  # 송신 Null여부
    ws.column_dimensions['E'].width = 20  # 수신 컬럼
    ws.column_dimensions['F'].width = 15  # 수신 타입
    ws.column_dimensions['G'].width = 12  # 수신 크기
    ws.column_dimensions['H'].width = 12  # 수신 Null여부
    ws.column_dimensions['I'].width = 40  # 비교 결과
    ws.column_dimensions['J'].width = 12  # 상태
    
    # 모든 셀에 테두리 적용
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = border

def auto_adjust_row_heights(ws):
    """행 높이를 자동으로 조정하는 함수"""
    for row in range(1, ws.max_row + 1):
        max_length = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                max_length = max(max_length, lines)
        # 기본 높이 20, 줄 수에 따라 높이 조정 (한 줄당 15픽셀)
        ws.row_dimensions[row].height = max(20, min(15 * max_length, 100))

def main():
    try:
        # input.xlsx 파일 로드
        input_xlsx_path = 'input.xlsx'
        wb_input = openpyxl.load_workbook(input_xlsx_path)
        ws_input = wb_input.active
        
        # output.xlsx 파일 생성
        output_xlsx_path = 'output.xlsx'
        wb_output = openpyxl.Workbook()
        # 기본 시트 제거
        wb_output.remove(wb_output.active)
        
        # 각 인터페이스 블록 처리
        interface_count = 0
        error_interfaces = []  # 오류가 발생한 인터페이스 정보를 저장할 리스트
        current_col = 2  # B열부터 시작
        
        while current_col <= ws_input.max_column:
            try:
                interface_info = read_interface_block(ws_input, current_col)
                if not interface_info:
                    break
                
                interface_count += 1
                interface_name = interface_info.get('interface_name', f'Interface_{interface_count}')
                
                print(f"\n처리 중인 인터페이스: {interface_name}")
                mapper = ColumnMapper()
                results = process_interface(interface_info, mapper)
                
                # 결과를 output.xlsx의 새로운 시트에 기록
                write_interface_result_to_sheet(wb_output, interface_info, results, interface_count)
                
                print(f"인터페이스 {interface_name} 처리 완료")
                
            except Exception as e:
                error_msg = f"인터페이스 처리 중 오류 발생 (열: {current_col}): {str(e)}"
                print(f"\n[오류] {error_msg}")
                error_interfaces.append({
                    'column': current_col,
                    'interface': interface_info.get('interface_name', f'Interface_{interface_count}') if interface_info else 'Unknown',
                    'error': str(e)
                })
            
            finally:
                current_col += 3  # 다음 인터페이스로 이동
        
        # 결과 파일 저장
        wb_output.save(output_xlsx_path)
        wb_input.close()
        wb_output.close()
        
        # 전체 처리 결과 출력
        print("\n=== 처리 완료 ===")
        print(f"총 처리된 인터페이스 수: {interface_count}")
        
        if error_interfaces:
            print("\n=== 오류 발생한 인터페이스 목록 ===")
            for error_info in error_interfaces:
                print(f"\n인터페이스: {error_info['interface']}")
                print(f"위치: 열 {error_info['column']}")
                print(f"오류 내용: {error_info['error']}")
        
    except Exception as e:
        print(f"\n[심각한 오류] 프로그램 실행 중 오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()
