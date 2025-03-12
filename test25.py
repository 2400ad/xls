import openpyxl
import ast

def safe_get_dict_value(dictionary, key, default=''):
    """
    딕셔너리에서 안전하게 값을 가져오는 유틸리티 함수
    
    Args:
        dictionary: 검색할 딕셔너리 (None일 수 있음)
        key: 검색할 키
        default: 기본값 (dictionary가 None이거나 key가 없을 경우 반환)
    
    Returns:
        찾은 값 또는 기본값
    """
    if dictionary is None:
        return default
    return dictionary.get(key, default)

def read_interface_block(ws, start_col):
    """Excel에서 3컬럼 단위로 하나의 인터페이스 정보를 읽습니다."""
    interface_info = {
        'interface_name': ws.cell(row=1, column=start_col).value or '',  # IF NAME (1행)
        'interface_id': ws.cell(row=2, column=start_col).value or '',    # IF ID (2행)
        'send': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None},
        'recv': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None}
    }
    
    try:
        # DB 연결 정보 (3행에서 읽기)
        send_db_info = ast.literal_eval(ws.cell(row=3, column=start_col).value)
        recv_db_info = ast.literal_eval(ws.cell(row=3, column=start_col + 1).value)
        interface_info['send']['db_info'] = send_db_info
        interface_info['recv']['db_info'] = recv_db_info
        
        # 테이블 정보 (4행에서 읽기)
        send_table_info = ast.literal_eval(ws.cell(row=4, column=start_col).value)
        recv_table_info = ast.literal_eval(ws.cell(row=4, column=start_col + 1).value)
        
        interface_info['send']['owner'] = send_table_info.get('owner')
        interface_info['send']['table_name'] = send_table_info.get('table_name')
        interface_info['recv']['owner'] = recv_table_info.get('owner')
        interface_info['recv']['table_name'] = recv_table_info.get('table_name')
        
        # 컬럼 매핑 정보 (5행부터)
        row = 5
        while True:
            send_col = ws.cell(row=row, column=start_col).value
            recv_col = ws.cell(row=row, column=start_col + 1).value
            
            if not send_col and not recv_col:
                break
                
            interface_info['send']['columns'].append(send_col if send_col else '')
            interface_info['recv']['columns'].append(recv_col if recv_col else '')
            row += 1
            
    except Exception as e:
        print(f'인터페이스 정보 읽기 중 오류 발생: {str(e)}')
        return None
    
    return interface_info

def read_all_interfaces(xlsx_path):
    """엑셀 파일에서 모든 인터페이스 정보를 읽어 딕셔너리로 반환합니다."""
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        interfaces = {}
        current_col = 2  # B열부터 시작
        
        while current_col <= ws.max_column:
            interface_info = read_interface_block(ws, current_col)
            if not interface_info:
                break
                
            # 인터페이스 ID를 키로 사용하여 저장
            interface_id = interface_info['interface_id'].strip()
            if interface_id:
                # 송신 컬럼과 수신 컬럼의 매핑 딕셔너리 생성
                send_recv_mapping = {}
                for i in range(len(interface_info['send']['columns'])):
                    send_col = interface_info['send']['columns'][i].strip() if interface_info['send']['columns'][i] else ''
                    recv_col = interface_info['recv']['columns'][i].strip() if interface_info['recv']['columns'][i] else ''
                    if send_col:  # 송신 컬럼이 있는 경우만 매핑
                        send_recv_mapping[send_col] = recv_col
                
                interfaces[interface_id] = {
                    'interface_name': interface_info['interface_name'],
                    'send_owner': interface_info['send']['owner'],
                    'send_table': interface_info['send']['table_name'],
                    'recv_owner': interface_info['recv']['owner'],
                    'recv_table': interface_info['recv']['table_name'],
                    'send_recv_mapping': send_recv_mapping
                }
            
            current_col += 3  # 다음 인터페이스로 이동
        
        wb.close()
        return interfaces
        
    except Exception as e:
        print(f'엑셀 파일 읽기 중 오류 발생: {str(e)}')
        return {}

def compare_interfaces(file1_interfaces, file2_interfaces):
    """두 엑셀 파일에서 읽은 인터페이스 정보를 비교합니다."""
    # 공통 인터페이스 ID 찾기
    common_interface_ids = set(file1_interfaces.keys()) & set(file2_interfaces.keys())
    
    # 파일1에만 있는 인터페이스 ID
    only_in_file1 = set(file1_interfaces.keys()) - set(file2_interfaces.keys())
    
    # 파일2에만 있는 인터페이스 ID
    only_in_file2 = set(file2_interfaces.keys()) - set(file1_interfaces.keys())
    
    comparison_results = {
        'common_interfaces': len(common_interface_ids),
        'only_in_file1': list(only_in_file1),
        'only_in_file2': list(only_in_file2),
        'interface_details': {},
        'mismatched_interfaces': []  # 불일치하는 인터페이스 목록 추가
    }
    
    # 각 공통 인터페이스에 대한 상세 비교
    for interface_id in common_interface_ids:
        file1_interface = file1_interfaces[interface_id]
        file2_interface = file2_interfaces[interface_id]
        
        # 송신 컬럼 비교 (공백 trim 후)
        file1_send_cols = set(file1_interface['send_recv_mapping'].keys())
        file2_send_cols = set(file2_interface['send_recv_mapping'].keys())
        
        # 파일1에만 있는 송신 컬럼
        only_in_file1_cols = file1_send_cols - file2_send_cols
        
        # 파일2에만 있는 송신 컬럼
        only_in_file2_cols = file2_send_cols - file1_send_cols
        
        # 공통 송신 컬럼
        common_send_cols = file1_send_cols & file2_send_cols
        
        # 수신 컬럼 비교
        recv_col_comparison = []
        mismatched_recv_cols = []  # 불일치하는 수신 컬럼 목록 추가
        
        for send_col in common_send_cols:
            file1_recv = file1_interface['send_recv_mapping'][send_col]
            file2_recv = file2_interface['send_recv_mapping'][send_col]
            
            # 수신 컬럼 이름 비교 (trim 후 정확히 일치하는지)
            is_match = file1_recv.strip() == file2_recv.strip()
            
            if not is_match:
                mismatched_recv_cols.append(send_col)
            
            recv_col_comparison.append({
                'send_col': send_col,
                'file1_recv': file1_recv,
                'file2_recv': file2_recv,
                'is_match': is_match
            })
        
        # 인터페이스 상세 비교 결과 저장
        all_cols_match = (len(only_in_file1_cols) == 0 and 
                         len(only_in_file2_cols) == 0 and 
                         len(mismatched_recv_cols) == 0)
        
        interface_detail = {
            'interface_name': file1_interface['interface_name'],
            'only_in_file1_cols': list(only_in_file1_cols),
            'only_in_file2_cols': list(only_in_file2_cols),
            'common_cols_count': len(common_send_cols),
            'recv_col_comparison': recv_col_comparison,
            'mismatched_recv_cols': mismatched_recv_cols,  # 불일치하는 수신 컬럼 목록 추가
            'all_cols_match': all_cols_match  # 모든 컬럼이 일치하는지 여부
        }
        
        comparison_results['interface_details'][interface_id] = interface_detail
        
        # 불일치하는 인터페이스가 있으면 목록에 추가
        if not all_cols_match:
            comparison_results['mismatched_interfaces'].append(interface_id)
    
    return comparison_results

def print_comparison_results(comparison_results, file1_interfaces, file2_interfaces):
    """비교 결과를 보기 좋게 출력합니다."""
    print("\n=== 인터페이스 비교 결과 요약 ===")
    print(f"공통 인터페이스 수: {comparison_results['common_interfaces']}")
    print(f"파일1에만 있는 인터페이스 수: {len(comparison_results['only_in_file1'])}")
    print(f"파일2에만 있는 인터페이스 수: {len(comparison_results['only_in_file2'])}")
    
    # 불일치하는 인터페이스 정보 추가
    mismatched_count = len(comparison_results['mismatched_interfaces'])
    print(f"불일치하는 인터페이스 수: {mismatched_count}")
    
    if mismatched_count > 0:
        print("\n[불일치하는 인터페이스 목록]")
        for idx, interface_id in enumerate(comparison_results['mismatched_interfaces']):
            interface_name = comparison_results['interface_details'][interface_id]['interface_name']
            print(f"{idx+1}. {interface_id} ({interface_name})")
            
            # 불일치 상세 정보 요약
            detail = comparison_results['interface_details'][interface_id]
            if detail['only_in_file1_cols']:
                print(f"   - 파일1에만 있는 송신 컬럼: {len(detail['only_in_file1_cols'])}개")
            if detail['only_in_file2_cols']:
                print(f"   - 파일2에만 있는 송신 컬럼: {len(detail['only_in_file2_cols'])}개")
            if detail['mismatched_recv_cols']:
                print(f"   - 수신 컬럼 불일치: {len(detail['mismatched_recv_cols'])}개")
    
    # 파일1에만 있는 인터페이스 목록
    if comparison_results['only_in_file1']:
        print("\n[파일1에만 있는 인터페이스 목록]")
        for idx, interface_id in enumerate(comparison_results['only_in_file1']):
            interface_name = file1_interfaces[interface_id]['interface_name']
            print(f"{idx+1}. {interface_id} ({interface_name})")
    
    # 파일2에만 있는 인터페이스 목록
    if comparison_results['only_in_file2']:
        print("\n[파일2에만 있는 인터페이스 목록]")
        for idx, interface_id in enumerate(comparison_results['only_in_file2']):
            interface_name = file2_interfaces[interface_id]['interface_name']
            print(f"{idx+1}. {interface_id} ({interface_name})")
    
    print("\n=== 인터페이스 상세 비교 ===")
    for interface_id, details in comparison_results['interface_details'].items():
        print(f"\n인터페이스 ID: {interface_id}")
        print(f"인터페이스 이름: {details['interface_name']}")
        print(f"공통 송신 컬럼 수: {details['common_cols_count']}")
        
        # 파일1에만 있는 송신 컬럼
        if details['only_in_file1_cols']:
            print(f"파일1에만 있는 송신 컬럼 ({len(details['only_in_file1_cols'])}개): {', '.join(details['only_in_file1_cols'])}")
        else:
            print("파일1에만 있는 송신 컬럼: 없음")
        
        # 파일2에만 있는 송신 컬럼
        if details['only_in_file2_cols']:
            print(f"파일2에만 있는 송신 컬럼 ({len(details['only_in_file2_cols'])}개): {', '.join(details['only_in_file2_cols'])}")
        else:
            print("파일2에만 있는 송신 컬럼: 없음")
        
        # 수신 컬럼 비교 표시
        if details['recv_col_comparison']:
            print("\n[송신-수신 컬럼 매핑 비교]")
            print("송신 컬럼 | 파일1 수신 컬럼 | 파일2 수신 컬럼 | 일치 여부")
            print("-" * 70)
            
            # 불일치하는 항목 먼저 표시
            for comp in details['recv_col_comparison']:
                if not comp['is_match']:
                    match_str = "불일치 ❌"
                    print(f"{comp['send_col']} | {comp['file1_recv']} | {comp['file2_recv']} | {match_str}")
            
            # 일치하는 항목 표시
            for comp in details['recv_col_comparison']:
                if comp['is_match']:
                    match_str = "일치 ✓"
                    print(f"{comp['send_col']} | {comp['file1_recv']} | {comp['file2_recv']} | {match_str}")
        
        # 전체 일치 여부 표시
        match_status = "모든 컬럼 일치 ✓" if details['all_cols_match'] else "불일치 항목 있음 ❌"
        print(f"\n[결과] {match_status}")

# 테스트 코드
if __name__ == "__main__":
    try:
        # 두 개의 엑셀 파일 경로
        input_xlsx_path1 = 'input1.xlsx'
        input_xlsx_path2 = 'input2.xlsx'
        
        print(f"파일1 '{input_xlsx_path1}'에서 인터페이스 정보 읽는 중...")
        file1_interfaces = read_all_interfaces(input_xlsx_path1)
        print(f"파일1에서 {len(file1_interfaces)}개의 인터페이스를 읽었습니다.")
        
        print(f"\n파일2 '{input_xlsx_path2}'에서 인터페이스 정보 읽는 중...")
        file2_interfaces = read_all_interfaces(input_xlsx_path2)
        print(f"파일2에서 {len(file2_interfaces)}개의 인터페이스를 읽었습니다.")
        
        print("\n두 파일의 인터페이스 비교 중...")
        comparison_results = compare_interfaces(file1_interfaces, file2_interfaces)
        
        # 비교 결과 출력
        if comparison_results is not None:
            print_comparison_results(comparison_results, file1_interfaces, file2_interfaces)
        else:
            print("비교 결과가 없습니다.")
        
    except Exception as e:
        print(f"\n오류 발생: {str(e)}")