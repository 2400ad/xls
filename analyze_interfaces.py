import openpyxl
import ast
from maptest import ColumnMapper

def read_interface_block(ws, start_col):
    """Excel에서 3컬럼 단위로 하나의 인터페이스 정보를 읽습니다.
    
    Args:
        ws: Worksheet
        start_col: 시작 컬럼 인덱스 (1부터 시작)
    
    Returns:
        interface_info: 인터페이스 정보를 담은 딕셔너리
    """
    # 첫번째 컬럼에서 정보 추출
    interface_name = ws.cell(row=1, column=start_col).value
    if not interface_name:  # 더 이상 인터페이스가 없음
        return None
        
    interface_id = ws.cell(row=2, column=start_col).value
    
    try:
        # 문자열로 된 dict를 실제 dict로 변환
        send_db = ast.literal_eval(ws.cell(row=3, column=start_col).value)
        send_table = ast.literal_eval(ws.cell(row=4, column=start_col).value)
        recv_db = ast.literal_eval(ws.cell(row=3, column=start_col+1).value)
        recv_table = ast.literal_eval(ws.cell(row=4, column=start_col+1).value)
    except Exception as e:
        print(f"Error parsing DB/Table info for interface {interface_name}: {str(e)}")
        return None
    
    # 컬럼 정보 시작 행
    start_row = 5
    
    # 빈 값이 아닌 행까지의 컬럼 정보 수집
    send_columns = []
    recv_columns = []
    comments = []
    
    max_row = ws.max_row
    for row in range(start_row, max_row + 1):
        send_val = ws.cell(row=row, column=start_col).value
        recv_val = ws.cell(row=row, column=start_col+1).value
        comment_val = ws.cell(row=row, column=start_col+2).value
        
        # 송신과 수신 컬럼이 모두 없는 경우에만 종료
        if not send_val and not recv_val:
            # 이전 행까지 모두 비어있는지 확인
            prev_send = ws.cell(row=row-1, column=start_col).value
            prev_recv = ws.cell(row=row-1, column=start_col+1).value
            if not prev_send and not prev_recv:
                break
            
        # 송신이나 수신 중 하나라도 있으면 추가
        if send_val is not None:  # 빈 문자열도 값으로 처리
            send_columns.append(send_val)
        if recv_val is not None:  # 빈 문자열도 값으로 처리
            recv_columns.append(recv_val)
        if comment_val:
            comments.append(comment_val)
    
    return {
        'name': interface_name,
        'id': interface_id,
        'send_db': send_db,
        'send_table': send_table,
        'recv_db': recv_db,
        'recv_table': recv_table,
        'send_columns': send_columns,
        'recv_columns': recv_columns,
        'comments': comments
    }

def analyze_excel():
    """input.xlsx 파일을 읽고 분석합니다."""
    try:
        # Excel 파일 읽기
        wb = openpyxl.load_workbook('input.xlsx', data_only=True)
        ws = wb.active
        
        # 3컬럼씩 인터페이스 정보 읽기 (B열부터 시작)
        interfaces = []
        for col in range(2, ws.max_column + 1, 3):  # 2는 B열을 의미
            interface_info = read_interface_block(ws, col)
            if interface_info:
                interfaces.append(interface_info)
            else:
                break  # 더 이상 인터페이스가 없음
        
        # 분석 결과 출력
        print(f"총 {len(interfaces)}개의 인터페이스를 발견했습니다.\n")
        
        for interface in interfaces:
            print("="*50)
            print(f"인터페이스 이름: {interface['name']}")
            print(f"인터페이스 ID: {interface['id']}")
            print("\n[송신 정보]")
            print(f"DB 정보: {interface['send_db']}")
            print(f"테이블 정보: {interface['send_table']}")
            print("컬럼 목록:")
            for i, col in enumerate(interface['send_columns']):
                print(f"  {i+1}. {col}")
            
            print("\n[수신 정보]")
            print(f"DB 정보: {interface['recv_db']}")
            print(f"테이블 정보: {interface['recv_table']}")
            print("컬럼 목록:")
            for i, col in enumerate(interface['recv_columns']):
                print(f"  {i+1}. {col}")
            
            print("\n[컬럼 매핑]")
            # 송신과 수신 컬럼을 별도로 출력
            print("송신 컬럼:")
            for i, col in enumerate(interface['send_columns']):
                print(f"  {i+1}. {col}")
                
            print("\n수신 컬럼:")
            for i, col in enumerate(interface['recv_columns']):
                print(f"  {i+1}. {col}")
                
            print("\n매핑 관계:")
            # 실제 매핑 관계만 출력
            mapping_count = min(len(interface['send_columns']), len(interface['recv_columns']))
            for i in range(mapping_count):
                comment = interface['comments'][i] if i < len(interface['comments']) else ""
                if interface['recv_columns'][i]:  # 수신 컬럼이 있는 경우만 매핑 표시
                    print(f"  {interface['send_columns'][i]} -> {interface['recv_columns'][i]}")
                    if comment:
                        print(f"    설명: {comment}")
            print()
            
    except Exception as e:
        print(f"Error analyzing Excel file: {str(e)}")
    finally:
        if 'wb' in locals():
            wb.close()

if __name__ == "__main__":
    analyze_excel()
