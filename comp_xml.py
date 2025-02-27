import sys
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List, Tuple, Optional
import xml.etree.ElementTree as ET
from comp_excel import ExcelManager, read_interface_block
from xltest import process_interface, read_interface_block
from comp_q import QueryParser, QueryDifference, FileSearcher, BWQueryExtractor
from maptest import ColumnMapper
import datetime
import ast

def read_interface_block(ws, start_col):
    """Excel에서 3컬럼 단위로 하나의 인터페이스 정보를 읽습니다.
    이 함수는 xltest.py의 동일한 함수를 대체하지 않고, 가져오지 못한 경우의 백업 역할만 합니다.
    """
    try:
        interface_info = {
            'interface_name': ws.cell(row=1, column=start_col).value or '',  # IF NAME (1행)
            'interface_id': ws.cell(row=2, column=start_col).value or '',    # IF ID (2행)
            'send': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None},
            'recv': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None}
        }
        
        # 인터페이스 ID가 없으면 빈 인터페이스로 간주
        if not interface_info['interface_id']:
            return None
            
        # DB 연결 정보 (3행에서 읽기)
        try:
            send_db_value = ws.cell(row=3, column=start_col).value
            send_db_info = ast.literal_eval(send_db_value) if send_db_value else {}
            
            recv_db_value = ws.cell(row=3, column=start_col + 1).value
            recv_db_info = ast.literal_eval(recv_db_value) if recv_db_value else {}
        except (SyntaxError, ValueError):
            # 데이터 형식 오류 시 빈 딕셔너리로 설정
            send_db_info = {}
            recv_db_info = {}
            
        interface_info['send']['db_info'] = send_db_info
        interface_info['recv']['db_info'] = recv_db_info
        
        # 테이블 정보 (4행에서 읽기)
        try:
            send_table_value = ws.cell(row=4, column=start_col).value
            send_table_info = ast.literal_eval(send_table_value) if send_table_value else {}
            
            recv_table_value = ws.cell(row=4, column=start_col + 1).value
            recv_table_info = ast.literal_eval(recv_table_value) if recv_table_value else {}
        except (SyntaxError, ValueError):
            # 데이터 형식 오류 시 빈 딕셔너리로 설정
            send_table_info = {}
            recv_table_info = {}
        
        interface_info['send']['owner'] = send_table_info.get('owner')
        interface_info['send']['table_name'] = send_table_info.get('table_name')
        interface_info['recv']['owner'] = recv_table_info.get('owner')
        interface_info['recv']['table_name'] = recv_table_info.get('table_name')
        
        # 컬럼 매핑 정보 (5행부터)
        row = 5
        while True:
            send_col = ws.cell(row=row, column=start_col).value
            recv_col = ws.cell(row=row, column=start_col + 1).value
            
            if not send_col and not recv_col:
                break
                
            interface_info['send']['columns'].append(send_col if send_col else '')
            interface_info['recv']['columns'].append(recv_col if recv_col else '')
            row += 1
            
    except Exception as e:
        print(f'인터페이스 정보 읽기 중 오류 발생: {str(e)}')
        return None
    
    return interface_info

class XMLComparator:
    # 클래스 변수로 BW_SEARCH_DIR 정의
    BW_SEARCH_DIR = "C:\\work\\LT\\BW소스"

    def __init__(self, excel_path: str, search_dir: str):
        """
        XML 비교를 위한 클래스 초기화
        
        Args:
            excel_path (str): 인터페이스 정보가 있는 Excel 파일 경로
            search_dir (str): XML 파일을 검색할 디렉토리 경로
        """
        self.excel_path = excel_path
        self.search_dir = search_dir
        self.workbook = openpyxl.load_workbook(excel_path)
        self.worksheet = self.workbook.active
        self.mapper = ColumnMapper()
        self.query_parser = QueryParser()
        self.excel_manager = ExcelManager()  # ExcelManager 인스턴스 생성
        self.interface_results = []  # 모든 인터페이스 처리 결과 저장
        self.output_path = 'C:\\work\\LT\\comp_mq_bw.xlsx'  # 기본 출력 경로

    def extract_from_xml(self, xml_path: str) -> Tuple[str, str]:
        """
        XML 파일에서 쿼리와 XML 내용을 추출합니다.
        
        Args:
            xml_path (str): XML 파일 경로
            
        Returns:
            Tuple[str, str]: (쿼리, XML 내용)
        """
        try:
            # XML 파일이 제대로 로드되었는지 확인
            if not os.path.exists(xml_path):
                print(f"Warning: XML file not found: {xml_path}")
                return None, None
                
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # XML 내용이 유효한지 확인
            if root is None:
                print(f"Warning: Invalid XML content in file: {xml_path}")
                return None, None
            
            # SQL 노드 찾기
            sql_node = root.find(".//SQL")
            if sql_node is None or not sql_node.text:
                print(f"Warning: No SQL content found in file: {xml_path}")
                return None, None
                
            query = sql_node.text.strip()
            xml_content = ET.tostring(root, encoding='unicode')
            
            # 추출된 쿼리가 유효한지 확인
            if not query:
                print(f"Warning: Empty SQL query in file: {xml_path}")
                return None, None
                
            return query, xml_content
            
        except ET.ParseError as e:
            print(f"Error parsing XML file {xml_path}: {e}")
            return None, None
        except Exception as e:
            print(f"Unexpected error processing file {xml_path}: {e}")
            return None, None
            
    def compare_queries(self, query1: str, query2: str) -> QueryDifference:
        """
        두 쿼리를 비교합니다.
        
        Args:
            query1 (str): 첫 번째 쿼리
            query2 (str): 두 번째 쿼리
            
        Returns:
            QueryDifference: 쿼리 비교 결과
        """
        if not query1 or not query2:
            return None
        return self.query_parser.compare_queries(query1, query2)
        
    def find_interface_files(self, if_id: str) -> Dict[str, Dict]:
        """
        주어진 IF ID에 해당하는 송수신 XML 파일을 찾고 쿼리를 추출합니다.
        파일명 패턴: {if_id}로 시작하고 .SND.xml 또는 .RCV.xml로 끝나는 파일
        
        Args:
            if_id (str): 인터페이스 ID
            
        Returns:
            Dict[str, Dict]: {
                'send': {'path': 송신파일경로, 'query': 송신쿼리, 'xml': 송신XML},
                'recv': {'path': 수신파일경로, 'query': 수신쿼리, 'xml': 수신XML}
            }
        """
        results = {
            'send': {'path': None, 'query': None, 'xml': None},
            'recv': {'path': None, 'query': None, 'xml': None}
        }
        
        if not if_id:
            print("Warning: Empty IF_ID provided")
            return results
            
        try:
            # 디렉토리 내의 모든 XML 파일 검색
            for file in os.listdir(self.search_dir):
                if not file.startswith(if_id):
                    continue
                    
                file_path = os.path.join(self.search_dir, file)
                
                # 송신 파일 (.SND.xml)
                if file.endswith('.SND.xml'):
                    results['send']['path'] = file_path
                    query, xml = self.extract_from_xml(file_path)
                    if query and xml:
                        results['send']['query'] = query
                        results['send']['xml'] = xml
                    else:
                        print(f"Warning: Failed to extract query from send file: {file_path}")
                
                # 수신 파일 (.RCV.xml)
                elif file.endswith('.RCV.xml'):
                    results['recv']['path'] = file_path
                    query, xml = self.extract_from_xml(file_path)
                    if query and xml:
                        results['recv']['query'] = query
                        results['recv']['xml'] = xml
                    else:
                        print(f"Warning: Failed to extract query from receive file: {file_path}")
            
            # 파일을 찾았는지 확인
            if not results['send']['path'] and not results['recv']['path']:
                print(f"Warning: No interface files found for IF_ID: {if_id}")
            elif not results['send']['path']:
                print(f"Warning: No send file found for IF_ID: {if_id}")
            elif not results['recv']['path']:
                print(f"Warning: No receive file found for IF_ID: {if_id}")
            
            return results
            
        except Exception as e:
            print(f"Error finding interface files: {e}")
            return results
        
    def process_interface_block(self, start_col: int) -> Optional[Dict]:
        """
        Excel에서 하나의 인터페이스 블록을 처리합니다.
        
        Args:
            start_col (int): 인터페이스 블록이 시작되는 컬럼
            
        Returns:
            Optional[Dict]: 처리된 인터페이스 정보와 결과, 실패시 None
        """
        try:
            # Excel에서 인터페이스 정보 읽기
            interface_info = read_interface_block(self.worksheet, start_col)
            if not interface_info:
                print(f"Warning: Failed to read interface block at column {start_col}")
                return None
                
            # Excel에서 추출된 쿼리와 XML 얻기
            excel_results = process_interface(interface_info, self.mapper)
            if not excel_results:
                print(f"Warning: Failed to process interface at column {start_col}")
                return None
                
            # 송수신 파일 찾기
            file_results = self.find_interface_files(interface_info['interface_id'])
            if not file_results:
                print(f"Warning: No interface files found for IF_ID: {interface_info['interface_id']}")
                return None
            
            # 결과 초기화
            comparisons = {
                'send': None,
                'recv': None
            }
            warnings = {
                'send': [],
                'recv': []
            }
            
            # 송신 쿼리 처리
            if excel_results['send_sql'] and file_results['send']['query']:
                try:
                    comparisons['send'] = self.query_parser.compare_queries(
                        excel_results['send_sql'],
                        file_results['send']['query']
                    )
                    warnings['send'].extend(
                        self.query_parser.check_special_columns(
                            file_results['send']['query'],
                            'send'
                        )
                    )
                except Exception as e:
                    print(f"Error comparing send queries: {e}")
                    print(f"Excel query: {excel_results['send_sql']}")
                    print(f"File query: {file_results['send']['query']}")
            
            # 수신 쿼리 처리
            if excel_results['recv_sql'] and file_results['recv']['query']:
                try:
                    comparisons['recv'] = self.query_parser.compare_queries(
                        excel_results['recv_sql'],
                        file_results['recv']['query']
                    )
                    warnings['recv'].extend(
                        self.query_parser.check_special_columns(
                            file_results['recv']['query'],
                            'recv'
                        )
                    )
                except Exception as e:
                    print(f"Error comparing receive queries: {e}")
                    print(f"Excel query: {excel_results['recv_sql']}")
                    print(f"File query: {file_results['recv']['query']}")
            
            return {
                'if_id': interface_info['interface_id'],
                'interface_name': interface_info['interface_name'],
                'comparisons': comparisons,
                'warnings': warnings,
                'excel': excel_results,
                'files': file_results
            }
            
        except Exception as e:
            print(f"Error processing interface block at column {start_col}: {e}")
            return None
            
    def process_all_interfaces(self) -> List[Dict]:
        """
        Excel 파일의 모든 인터페이스를 처리합니다.
        B열부터 시작하여 3컬럼 단위로 처리합니다.
        
        Returns:
            List[Dict]: 각 인터페이스의 처리 결과 목록
        """
        results = []
        col = 2  # B열부터 시작
        
        while True:
            # 인터페이스 ID가 없으면 종료
            if not self.worksheet.cell(row=2, column=col).value:
                break
                
            result = self.process_interface_block(col)
            if result:
                results.append(result)
                
            col += 3  # 다음 인터페이스 블록으로 이동
            
        # 결과 출력
        for idx, result in enumerate(results, 1):
            print(f"\n=== 인터페이스 {idx} ===")
            print(f"ID: {result['if_id']}")
            print(f"이름: {result['interface_name']}")
            
            print("\n파일 검색 결과:")
            print(f"송신 파일: {result['files']['send']['path']}")
            print(f"수신 파일: {result['files']['recv']['path']}")
            
            print("\n쿼리 비교 결과:")
            if result['comparisons']['send']:
                print("송신 쿼리:")
                print(f"  {result['comparisons']['send']}")
            if result['comparisons']['recv']:
                print("수신 쿼리:")
                print(f"  {result['comparisons']['recv']}")
            
            # 경고가 있을 때만 경고 섹션 출력
            send_warnings = result['warnings']['send']
            recv_warnings = result['warnings']['recv']
            if send_warnings or recv_warnings:
                print("\n경고:")
                if send_warnings:
                    print("송신 쿼리 경고:")
                    for warning in send_warnings:
                        print(f"  - {warning}")
                if recv_warnings:
                    print("수신 쿼리 경고:")
                    for warning in recv_warnings:
                        print(f"  - {warning}")
                
        return results
        
    def close(self):
        """리소스 정리"""
        self.workbook.close()
        if self.mapper:
            self.mapper.close_connections()

    def find_bw_files(self) -> List[Dict[str, str]]:
        """
        엑셀의 인터페이스 정보에서 송신 테이블명을 추출하여 BW 파일을 검색합니다.
        
        Returns:
            List[Dict[str, str]]: [
                {
                    'interface_name': str,
                    'interface_id': str,
                    'send_table': str,
                    'bw_files': List[str]
                },
                ...
            ]
        """
        results = []
        file_searcher = FileSearcher()
        
        # 엑셀에서 인터페이스 정보 읽기
        for row in range(2, self.worksheet.max_row + 1, 3):  # 3행씩 건너뛰며 읽기
            interface_info = read_interface_block(self.worksheet, row)
            if not interface_info:
                continue
                
            # 송신 테이블명 추출 (스키마/오너 제외)
            send_table = interface_info['send'].get('table_name')
            if not send_table:
                continue
                
            # BW 파일 검색 - self.BW_SEARCH_DIR 사용
            bw_files = file_searcher.find_files_with_keywords(self.BW_SEARCH_DIR, [send_table])
            matching_files = bw_files.get(send_table, [])
            
            results.append({
                'interface_name': interface_info['interface_name'],
                'interface_id': interface_info['interface_id'],
                'send_table': send_table,
                'bw_files': matching_files
            })
            
        return results
        
    def print_bw_search_results(self, results: List[Dict[str, str]]):
        """
        BW 파일 검색 결과를 출력합니다.
        
        Args:
            results (List[Dict[str, str]]): find_bw_files()의 반환값
        """
        print("\nBW File Search Results:")
        print("-" * 80)
        print(f"{'Interface Name':<30} {'Interface ID':<15} {'Send Table':<20} {'BW Files'}")
        print("-" * 80)
        
        for result in results:
            bw_files_str = ', '.join(result['bw_files']) if result['bw_files'] else 'No matching files'
            print(f"{result['interface_name']:<30} {result['interface_id']:<15} {result['send_table']:<20} {bw_files_str}")

    def initialize_excel_output(self):
        """
        결과를 저장할 새 엑셀 파일 초기화
        """
        # ExcelManager를 통해 Excel 출력을 초기화
        self.excel_manager.initialize_excel_output()
        
    def save_excel_output(self, output_path=None):
        """
        처리된 결과를 엑셀 파일로 저장
        
        Args:
            output_path (str, optional): 출력 엑셀 파일 경로, 없으면 기본 경로 사용
            
        Returns:
            bool: 저장 성공 여부
        """
        # output_path 값을 인스턴스 변수에 저장
        if output_path:
            self.output_path = output_path
            
        # ExcelManager를 사용하여 파일 저장
        return self.excel_manager.save_excel_output(self.output_path)
        
    def create_interface_sheet(self, if_info, file_results, query_comparisons, bw_queries=None, bw_files=None):
        """
        인터페이스 정보와 비교 결과를 포함하는 엑셀 시트를 생성합니다.
        
        Args:
            if_info (dict): 인터페이스 정보
            file_results (dict): MQ 파일 결과 (송신/수신)
            query_comparisons (dict): 쿼리 비교 결과 (송신/수신)
            bw_queries (dict, optional): BW 쿼리 정보. Defaults to None.
            bw_files (list, optional): BW 매핑 파일 목록. Defaults to None.
        """
        # 기본값 설정
        bw_queries = bw_queries or {'send': '', 'recv': ''}
        bw_files = bw_files or []
        
        # 인터페이스 ID와 이름 확인
        if 'interface_id' not in if_info or not if_info['interface_id']:
            print("인터페이스 ID가 없습니다.")
            return
            
        # BW 파일 매핑
        bw_files_dict = {
            'send': bw_files[0] if bw_files and len(bw_files) > 0 else 'N/A',
            'recv': bw_files[1] if bw_files and len(bw_files) > 1 else 'N/A'
        }
        
        # MQ 파일 정보
        mq_files = {
            'send': file_results.get('send', {}),
            'recv': file_results.get('recv', {})
        }
        
        # 쿼리 정보 구성
        queries = {
            'mq_send': file_results.get('send', {}).get('query', 'N/A'),
            'bw_send': bw_queries.get('send', 'N/A'),
            'mq_recv': file_results.get('recv', {}).get('query', 'N/A'),
            'bw_recv': bw_queries.get('recv', 'N/A')
        }
        
        # 비교 결과 구성
        comparison_results = {
            'send': {
                'is_equal': query_comparisons.get('send', QueryDifference()).is_equal,
                'detail': query_comparisons.get('send', QueryDifference()).detail
            },
            'recv': {
                'is_equal': query_comparisons.get('recv', QueryDifference()).is_equal,
                'detail': query_comparisons.get('recv', QueryDifference()).detail
            }
        }
        
        # 인터페이스 시트 생성
        self.excel_manager.create_interface_sheet(if_info, mq_files, bw_files_dict, queries, comparison_results)
        
    def process_interface_with_bw(self, start_col: int, interface_info: Dict) -> Optional[Dict]:
        """
        하나의 인터페이스를 처리하고 BW 파일과 비교하여 결과 반환
        
        Args:
            start_col (int): 인터페이스 블록이 시작되는 컬럼
            interface_info (Dict): 인터페이스 정보
            
        Returns:
            Optional[Dict]: 처리된 인터페이스 정보와 결과, 실패시 None
        """
        try:
            # 표준 필드 생성
            # DB 정보에서 시스템 정보 추출
            if 'send' in interface_info and 'db_info' in interface_info['send'] and interface_info['send']['db_info']:
                interface_info['send_system'] = interface_info['send']['db_info'].get('system', 'N/A')
            else:
                interface_info['send_system'] = 'N/A'
                
            if 'recv' in interface_info and 'db_info' in interface_info['recv'] and interface_info['recv']['db_info']:
                interface_info['recv_system'] = interface_info['recv']['db_info'].get('system', 'N/A')
            else:
                interface_info['recv_system'] = 'N/A'
                
            # 테이블 정보 추출
            if 'send' in interface_info and 'table_name' in interface_info['send']:
                interface_info['send_table'] = interface_info['send']['table_name']
            else:
                interface_info['send_table'] = ''
                
            if 'recv' in interface_info and 'table_name' in interface_info['recv']:
                interface_info['recv_table'] = interface_info['recv']['table_name']
            else:
                interface_info['recv_table'] = ''
                
            # Excel에서 추출된 쿼리와 XML 얻기
            excel_results = process_interface(interface_info, self.mapper)
            if not excel_results:
                print(f"Warning: Failed to process interface at column {start_col}")
                return None
                
            # 송수신 파일 찾기
            file_results = self.find_interface_files(interface_info['interface_id'])
            if not file_results:
                print(f"Warning: No interface files found for IF_ID: {interface_info['interface_id']}")
                return None
            
            # BW 파일 찾기
            send_table = interface_info.get('send_table', '')
            if not send_table:
                print(f"Warning: No send table information for IF_ID: {interface_info['interface_id']}")
                bw_files = []
            else:
                # 송신 테이블로 BW 파일 검색
                bw_searcher = FileSearcher()
                bw_files = bw_searcher.find_files_with_keywords(
                    self.BW_SEARCH_DIR, 
                    [send_table]
                )
            
            # bw_files가 사전 형태이므로 send_table 키워드에 대한 결과를 가져옴
            matching_files = bw_files.get(send_table, [])
            
            # BW 쿼리 추출
            bw_queries = {
                'send': '',
                'recv': ''
            }
            extractor = BWQueryExtractor()
            for bw_file in matching_files:
                bw_file_path = os.path.join(self.BW_SEARCH_DIR, bw_file)
                if os.path.exists(bw_file_path):
                    # BWQueryExtractor의 extract_bw_queries 메서드를 사용하여 송신/수신 쿼리 모두 추출
                    queries = extractor.extract_bw_queries(bw_file_path)
                    
                    # 송신 쿼리가 없으면 첫 번째 송신 쿼리 저장
                    if not bw_queries['send'] and queries.get('send') and len(queries['send']) > 0:
                        bw_queries['send'] = queries['send'][0]
                    
                    # 수신 쿼리가 없으면 첫 번째 수신 쿼리 저장
                    if not bw_queries['recv'] and queries.get('recv') and len(queries['recv']) > 0:
                        bw_queries['recv'] = queries['recv'][0]
            
            # 결과 초기화
            comparisons = {
                'send': None,
                'recv': None
            }
            warnings = {
                'send': [],
                'recv': []
            }
            
            # 송신 쿼리 비교 (MQ XML vs BW XML)
            if file_results['send']['query'] and bw_queries['send']:
                try:
                    comparisons['send'] = self.query_parser.compare_queries(
                        file_results['send']['query'],
                        bw_queries['send']
                    )
                    warnings['send'].extend(
                        self.query_parser.check_special_columns(
                            file_results['send']['query'],
                            'send'
                        )
                    )
                except Exception as e:
                    print(f"Error comparing send queries: {e}")
                    print(f"MQ query: {file_results['send']['query']}")
                    print(f"BW query: {bw_queries['send']}")
            
            # 수신 쿼리 비교 (MQ XML vs BW XML)
            if file_results['recv']['query'] and bw_queries['recv']:
                try:
                    comparisons['recv'] = self.query_parser.compare_queries(
                        file_results['recv']['query'],
                        bw_queries['recv']
                    )
                    warnings['recv'].extend(
                        self.query_parser.check_special_columns(
                            file_results['recv']['query'],
                            'recv'
                        )
                    )
                except Exception as e:
                    print(f"Error comparing recv queries: {e}")
                    print(f"MQ query: {file_results['recv']['query']}")
                    print(f"BW query: {bw_queries['recv']}")
            
            # 결과 반환
            return {
                'interface_info': interface_info,
                'excel_results': excel_results,
                'file_results': file_results,
                'bw_queries': bw_queries,
                'comparisons': comparisons,
                'warnings': warnings,
                'bw_files': matching_files
            }
            
        except Exception as e:
            print(f"Error processing interface at column {start_col}: {e}")
            import traceback
            traceback.print_exc()
            return None

    def process_all_interfaces_with_bw(self):
        """
        모든 인터페이스를 처리하고 BW 파일과 비교하여 엑셀 파일로 결과 저장
        """
        # 엑셀 파일 초기화 - ExcelManager 사용
        self.excel_manager.initialize_excel_output()
        
        # 모든 열을 처리
        print("\n[인터페이스 처리 시작]")
        print("-" * 80)
        
        interface_count = 0
        processed_count = 0
        
        start_col = 2
        while True:
            interface_info = read_interface_block(self.worksheet, start_col)
            
            if not interface_info:
                break
                
            interface_count += 1
            
            # 인터페이스 ID와 이름 출력
            print(f"처리 중: [{interface_count}] {interface_info['interface_id']} - {interface_info['interface_name']}")
            
            # 인터페이스 처리 및 BW 비교
            result = self.process_interface_with_bw(start_col, interface_info)
            
            # 다음 인터페이스로 이동 (3칸씩)
            start_col += 3
            
            # 인터페이스 처리 결과가 있으면 엑셀에 저장
            if result:
                processed_count += 1
                
                # 결과를 저장할 인터페이스 시트 생성
                if_info = result['interface_info']
                
                # ExcelManager를 사용하여 인터페이스 시트 생성
                # MQ 파일 정보
                mq_files = {
                    'send': result['file_results']['send'],
                    'recv': result['file_results']['recv']
                }
                
                # BW 파일 정보
                bw_files = {
                    'send': result.get('bw_files', [])[0] if result.get('bw_files') and len(result.get('bw_files')) > 0 else 'N/A',
                    'recv': result.get('bw_files', [])[1] if result.get('bw_files') and len(result.get('bw_files')) > 1 else 'N/A'
                }
                
                # 쿼리 정보
                queries = {
                    'mq_send': result['file_results']['send']['query'],
                    'bw_send': result['bw_queries']['send'],
                    'mq_recv': result['file_results']['recv']['query'],
                    'bw_recv': result['bw_queries']['recv']
                }
                
                # 비교 결과
                comparison_results = {
                    'send': {
                        'is_equal': result['comparisons']['send'].is_equal if result['comparisons']['send'] else False,
                        'detail': result['comparisons']['send'].detail if result['comparisons']['send'] else '비교 불가'
                    },
                    'recv': {
                        'is_equal': result['comparisons']['recv'].is_equal if result['comparisons']['recv'] else False,
                        'detail': result['comparisons']['recv'].detail if result['comparisons']['recv'] else '비교 불가'
                    }
                }
                
                self.excel_manager.create_interface_sheet(
                    if_info, 
                    mq_files, 
                    bw_files, 
                    queries, 
                    comparison_results
                )
                
                # 요약 시트 업데이트
                self.update_summary_sheet(result, interface_count + 1)
        
        # 결과 저장
        self.save_excel_output()
        
        # 처리 결과 출력
        print("\n" + "=" * 80)
        print(f"처리 완료: 총 {interface_count}개 인터페이스 중 {processed_count}개 처리됨")
        print(f"결과 파일: {self.output_path}")
        print("=" * 80)
        
    def update_summary_sheet(self, result, row):
        """
        요약 시트에 현재 인터페이스 처리 결과를 추가합니다.
        
        Args:
            result (dict): 인터페이스 처리 결과
            row (int): 추가할 행 번호
        """
        # ExcelManager를 사용하여 요약 시트 업데이트
        self.excel_manager.update_summary_sheet(result, row)

    def extract_bw_queries(self, bw_results):
        """
        BW 파일에서 쿼리를 추출합니다.
        
        Args:
            bw_results (list): BW 파일 검색 결과 목록
            
        Returns:
            list: 인터페이스별 BW 쿼리 정보가 담긴 리스트
        """
        extractor = BWQueryExtractor()
        results = []
        
        for result in bw_results:
            if result['bw_files']:  # BW 파일이 있는 경우에만 처리
                print(f"\n인터페이스: {result['interface_name']} ({result['interface_id']})")
                print(f"송신 테이블: {result['send_table']}")
                print("찾은 BW 파일의 쿼리:")
                
                bw_queries = {'send': '', 'recv': ''}
                
                for bw_file in result['bw_files']:
                    bw_file_path = os.path.join(self.BW_SEARCH_DIR, bw_file)
                    if os.path.exists(bw_file_path):
                        # BWQueryExtractor의 extract_bw_queries 메서드를 사용하여 송신/수신 쿼리 모두 추출
                        queries = extractor.extract_bw_queries(bw_file_path)
                        
                        if queries['send'] and not bw_queries['send']:
                            bw_queries['send'] = queries['send'][0] if queries['send'] else ''
                            print(f"\nBW 송신 파일: {bw_file}")
                            print("-" * 40)
                            print(bw_queries['send'])
                            
                        if queries['recv'] and not bw_queries['recv']:
                            bw_queries['recv'] = queries['recv'][0] if queries['recv'] else ''
                            print(f"\nBW 수신 파일: {bw_file}")
                            print("-" * 40)
                            print(bw_queries['recv'])
                
                # 인터페이스 결과에 BW 쿼리 추가
                for interface_result in self.interface_results:
                    if interface_result['interface_info']['interface_id'] == result['interface_id']:
                        interface_result['bw_queries'] = bw_queries
                        interface_result['bw_files'] = result['bw_files']
                        break
                
                results.append({
                    'interface_id': result['interface_id'],
                    'bw_queries': bw_queries,
                    'bw_files': result['bw_files']
                })
        
        return results

def main():
    # 고정된 경로 사용
    excel_path = 'C:\\work\\LT\\input_LT.xlsx' # 인터페이스 정보
    xml_dir = 'C:\\work\\LT\\xml' # MQ XML 파일 디렉토리
    bw_dir = 'C:\\work\\LT\\BW소스'  # BW XML 파일 디렉토리 경로
    output_path = 'C:\\work\\LT\\comp_mq_bw.xlsx'  # 출력 엑셀 파일 경로
    
    # BW 검색 디렉토리 설정
    XMLComparator.BW_SEARCH_DIR = bw_dir
    
    # XML 비교기 초기화
    comparator = XMLComparator(excel_path, xml_dir)
    
    # 명령행 인자가 있을 경우 처리
    if len(sys.argv) > 1:
        if sys.argv[1] == "excel":
            # 엑셀 출력 모드 실행
            print("\n[MQ XML과 BW XML 쿼리 비교 - 엑셀 출력 모드]")
            comparator.process_all_interfaces_with_bw()
            return
        elif len(sys.argv) > 2 and sys.argv[1] == "output":
            # 출력 경로 변경
            output_path = sys.argv[2]
            comparator.output_path = output_path
            print(f"\n[출력 경로 변경: {output_path}]")
    
    # 기본 모드 실행 - 기존 로직 유지
    print("\n[MQ XML 파일 검색 및 쿼리 비교 시작]")
    comparator.process_all_interfaces()
    
    # BW 파일 검색 및 결과 출력을 마지막으로 이동
    print("\n[BW 파일 검색 시작]")
    bw_results = comparator.find_bw_files()
    comparator.print_bw_search_results(bw_results)
    
    # BW 파일에서 쿼리 추출
    print("\n[BW 파일 쿼리 추출]")
    print("-" * 80)
    bw_queries = comparator.extract_bw_queries(bw_results)
    
    # 처리 결과를 Excel로 저장 (excel 모드가 아닌 경우)
    print("\n[결과를 Excel로 저장]")
    comparator.initialize_excel_output()
    
    # 인터페이스별 결과 처리
    for i, result in enumerate(comparator.interface_results):
        if_info = result['interface_info']
        
        # 인터페이스 시트 생성
        comparator.create_interface_sheet(
            if_info, 
            result['file_results'], 
            result['comparisons'],
            result.get('bw_queries', {'send': '', 'recv': ''}),
            result.get('bw_files', [])
        )
        
        # 요약 시트 업데이트
        comparator.update_summary_sheet(result, i + 2)
    
    # 결과 저장
    comparator.save_excel_output(output_path)
    print(f"\n[분석 완료] 결과가 저장되었습니다: {output_path}")
    
    print("\n[처리 완료]")
    print("엑셀 출력 모드로 실행하려면 'python comp_xml.py excel' 명령을 사용하세요.")

if __name__ == "__main__":
    main()