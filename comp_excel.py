import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from typing import Dict, List, Optional, Tuple
import datetime
import ast


def read_interface_block(ws, start_col):
    """Excel?먯꽌 3而щ읆 ?⑥쐞濡??섎굹???명꽣?섏씠???뺣낫瑜??쎌뒿?덈떎.
    """
    try:
        interface_info = {
            'interface_name': ws.cell(row=1, column=start_col).value or '',  # IF NAME (1??
            'interface_id': ws.cell(row=2, column=start_col).value or '',    # IF ID (2??
            'send': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None},
            'recv': {'owner': None, 'table_name': None, 'columns': [], 'db_info': None}
        }
        
        # ?명꽣?섏씠??ID媛 ?놁쑝硫?鍮??명꽣?섏씠?ㅻ줈 媛꾩＜
        if not interface_info['interface_id']:
            return None
            
        # DB ?곌껐 ?뺣낫 (3?됱뿉???쎄린)
        try:
            send_db_value = ws.cell(row=3, column=start_col).value
            send_db_info = ast.literal_eval(send_db_value) if send_db_value else {}
            
            recv_db_value = ws.cell(row=3, column=start_col + 1).value
            recv_db_info = ast.literal_eval(recv_db_value) if recv_db_value else {}
        except (SyntaxError, ValueError):
            # ?곗씠???뺤떇 ?ㅻ쪟 ??鍮??뺤뀛?덈━濡??ㅼ젙
            send_db_info = {}
            recv_db_info = {}
            
        interface_info['send']['db_info'] = send_db_info
        interface_info['recv']['db_info'] = recv_db_info
        
        # ?뚯씠釉??뺣낫 (4?됱뿉???쎄린)
        try:
            send_table_value = ws.cell(row=4, column=start_col).value
            send_table_info = ast.literal_eval(send_table_value) if send_table_value else {}
            
            recv_table_value = ws.cell(row=4, column=start_col + 1).value
            recv_table_info = ast.literal_eval(recv_table_value) if recv_table_value else {}
        except (SyntaxError, ValueError):
            # ?곗씠???뺤떇 ?ㅻ쪟 ??鍮??뺤뀛?덈━濡??ㅼ젙
            send_table_info = {}
            recv_table_info = {}
        
        interface_info['send']['owner'] = send_table_info.get('owner')
        interface_info['send']['table_name'] = send_table_info.get('table_name')
        interface_info['recv']['owner'] = recv_table_info.get('owner')
        interface_info['recv']['table_name'] = recv_table_info.get('table_name')
        
        # 而щ읆 留ㅽ븨 ?뺣낫 (5?됰???
        row = 5
        while True:
            send_col = ws.cell(row=row, column=start_col).value
            recv_col = ws.cell(row=row, column=start_col + 1).value
            
            # ????None?대㈃ 而щ읆 留ㅽ븨 ??            if send_col is None and recv_col is None:
                break
                
            # ?≪떊 而щ읆 異붽?
            if send_col:
                interface_info['send']['columns'].append(send_col)
                
            # ?섏떊 而щ읆 異붽?
            if recv_col:
                interface_info['recv']['columns'].append(recv_col)
                
            row += 1
        
        return interface_info
        
    except Exception as e:
        print(f"?명꽣?섏씠???뺣낫 ?쎄린 ?ㅻ쪟: {e}")
        return None


class ExcelManager:
    """
    Excel ?뚯씪 愿由?諛?異쒕젰???꾪븳 ?대옒??    """

    def __init__(self, excel_path: str = None):
        """
        Excel 愿由ъ옄 ?대옒??珥덇린??        
        Args:
            excel_path (str, optional): 湲곗〈 ?묒? ?뚯씪 寃쎈줈 (?놁쑝硫??덈줈 ?앹꽦)
        """
        self.excel_path = excel_path
        if excel_path and os.path.exists(excel_path):
            self.workbook = openpyxl.load_workbook(excel_path)
        else:
            self.workbook = openpyxl.Workbook()
        
        self.output_path = ''
        
        # 寃곌낵 ?곹깭???곕Ⅸ ?됱긽 ?뺤쓽
        self.match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
        self.mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??        self.unavailable_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???
    def initialize_excel_output(self):
        """
        寃곌낵瑜???ν븷 ???묒? ?뚯씪 珥덇린??        
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: ?앹꽦???쒗듃 媛앹껜
        """
        # 湲곗〈 ?쒗듃媛 ?덉쑝硫?紐⑤몢 ??젣
        for sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        
        # ?붿빟 ?쒗듃 ?앹꽦
        sheet = self.workbook.create_sheet("?붿빟")
        
        # ?쒕ぉ ???ㅽ????뺤쓽
        header_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        wrap_text_top = Alignment(wrap_text=True, vertical='top')
        
        # ???덈퉬 ?ㅼ젙 - 紐낆떆?곸쑝濡??ㅼ젙
        sheet.column_dimensions['A'].width = 5   # ?쇰젴踰덊샇
        sheet.column_dimensions['B'].width = 20  # ?명꽣?섏씠??ID
        sheet.column_dimensions['C'].width = 25  # ?명꽣?섏씠??紐?        sheet.column_dimensions['D'].width = 25  # ?≪떊 ?뚯씠釉?        sheet.column_dimensions['E'].width = 25  # MQ ?≪떊 ?뚯씪
        sheet.column_dimensions['F'].width = 25  # BW ?≪떊 ?뚯씪
        sheet.column_dimensions['G'].width = 15  # ?≪떊 鍮꾧탳 寃곌낵
        sheet.column_dimensions['H'].width = 25  # MQ ?섏떊 ?뚯씪
        sheet.column_dimensions['I'].width = 25  # BW ?섏떊 ?뚯씪
        sheet.column_dimensions['J'].width = 15  # ?섏떊 鍮꾧탳 寃곌낵
        
        # ?ㅻ뜑 ???앹꽦
        headers = ["踰덊샇", "?명꽣?섏씠??ID", "?명꽣?섏씠??紐?, "?≪떊 ?뚯씠釉?, "MQ ?≪떊 ?뚯씪", "BW ?≪떊 ?뚯씪", "?≪떊 鍮꾧탳 寃곌낵", 
                  "MQ ?섏떊 ?뚯씪", "BW ?섏떊 ?뚯씪", "?섏떊 鍮꾧탳 寃곌낵"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=9)  # 湲瑗??ш린瑜?9濡??ㅼ젙
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border
        
        return sheet

    def update_summary_sheet(self, data, row=None):
        """
        ?붿빟 ?쒗듃???명꽣?섏씠???뺣낫 異붽?
        
        Args:
            data (dict): ?명꽣?섏씠???곗씠??            row (int, optional): 異붽?????踰덊샇 (None?대㈃ 留덉?留????ㅼ쓬??異붽?)
        """
        sheet = self.workbook["?붿빟"] if "?붿빟" in self.workbook.sheetnames else self.workbook.active
        
        # 異붽?????踰덊샇 寃곗젙
        if row is None:
            # 留덉?留????ㅼ쓬??異붽?
            row = 2
            for r in range(2, sheet.max_row + 2):
                if sheet.cell(row=r, column=2).value is None:
                    row = r
                    break
        
        # ?쇰젴踰덊샇 怨꾩궛
        seq_num = row - 1
        seq_num_formatted = f"{seq_num:02d}"  # 01, 02, ... ?뺤떇?쇰줈 ?щ㎎??        
        # ?명꽣?섏씠???뺣낫
        interface_info = data.get("interface_info", {})
        file_results = data.get("file_results", {})
        bw_files = data.get("bw_files", [])
        comparisons = data.get("comparisons", {})
        
        # 紐⑤뱺 ???湲곕낯 湲瑗??ш린 ?ㅼ젙
        font_size = 9
        font_normal = Font(size=font_size)
        
        # 媛??ㅼ젙
        cell = sheet.cell(row=row, column=1, value=seq_num_formatted)  # ?쇰젴踰덊샇
        cell.font = font_normal
        
        cell = sheet.cell(row=row, column=2, value=interface_info.get("interface_id", ""))
        cell.font = font_normal
        
        cell = sheet.cell(row=row, column=3, value=interface_info.get("interface_name", ""))
        cell.font = font_normal
        
        cell = sheet.cell(row=row, column=4, value=f"{interface_info.get('send', {}).get('owner', '')}.{interface_info.get('send', {}).get('table_name', '')}")
        cell.font = font_normal
        
        # MQ ?뚯씪 ?뺣낫
        cell = sheet.cell(row=row, column=5, value=file_results.get("send", {}).get("path", ""))
        cell.font = font_normal
        
        # BW ?뚯씪 ?뺣낫
        if isinstance(bw_files, list) and len(bw_files) > 0:
            cell = sheet.cell(row=row, column=6, value=bw_files[0])
        elif isinstance(bw_files, dict):
            cell = sheet.cell(row=row, column=6, value=bw_files.get("send", ""))
        cell.font = font_normal
        
        # 鍮꾧탳 寃곌낵 - ?≪떊
        send_comparison = comparisons.get("send", {})
        # 臾몄옄?댁씤 寃쎌슦 癒쇱? ?뺤씤
        if isinstance(send_comparison, str):
            # "?쇱튂" ?먮뒗 "遺덉씪移? 臾몄옄??洹몃?濡??ъ슜
            cell = sheet.cell(row=row, column=7, value=send_comparison)
            # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
            if send_comparison == "?쇱튂":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
            elif send_comparison == "遺덉씪移?:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            else:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        # QueryDifference ?대옒???몄뒪?댁뒪??寃쎌슦
        elif hasattr(send_comparison, "is_equal") and not isinstance(send_comparison, dict):
            # QueryDifference ?대옒?ㅼ쓽 __str__ 硫붿꽌???ъ슜?섏뿬 "?쇱튂" ?먮뒗 "遺덉씪移? 媛?몄삤湲?            result_str = str(send_comparison)
            cell = sheet.cell(row=row, column=7, value=result_str)
            # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
            if result_str == "?쇱튂":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
            elif result_str == "遺덉씪移?:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            else:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        # ?뺤뀛?덈━??寃쎌슦 ?뺤씤
        elif isinstance(send_comparison, dict):
            if "is_equal" in send_comparison:
                is_equal = send_comparison.get("is_equal", False)
                result_str = "?쇱튂" if is_equal else "遺덉씪移?
                cell = sheet.cell(row=row, column=7, value=result_str)
                # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
                if is_equal:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            # ?먮뒗 "detail" ?꾨뱶媛 ?덉쓣 寃쎌슦
            elif "detail" in send_comparison:
                detail = send_comparison.get("detail", "")
                if detail == "?쇱튂" or detail == "遺덉씪移?:
                    cell = sheet.cell(row=row, column=7, value=detail)
                    # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
                    if detail == "?쇱튂":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??                else:
                    cell = sheet.cell(row=row, column=7, value="遺덉씪移?)
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            else:
                cell = sheet.cell(row=row, column=7, value="鍮꾧탳遺덇?")
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        else:
            cell = sheet.cell(row=row, column=7, value="鍮꾧탳遺덇?")
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        cell.font = font_normal
        
        # MQ ?섏떊 ?뚯씪
        cell = sheet.cell(row=row, column=8, value=file_results.get("recv", {}).get("path", ""))
        cell.font = font_normal
        
        # BW ?섏떊 ?뚯씪
        if isinstance(bw_files, list) and len(bw_files) > 1:
            cell = sheet.cell(row=row, column=9, value=bw_files[1])
        elif isinstance(bw_files, dict):
            cell = sheet.cell(row=row, column=9, value=bw_files.get("recv", ""))
        cell.font = font_normal
        
        # 鍮꾧탳 寃곌낵 - ?섏떊
        recv_comparison = comparisons.get("recv", {})
        # 臾몄옄?댁씤 寃쎌슦 癒쇱? ?뺤씤
        if isinstance(recv_comparison, str):
            # "?쇱튂" ?먮뒗 "遺덉씪移? 臾몄옄??洹몃?濡??ъ슜
            cell = sheet.cell(row=row, column=10, value=recv_comparison)
            # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
            if recv_comparison == "?쇱튂":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
            elif recv_comparison == "遺덉씪移?:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            else:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        # QueryDifference ?대옒???몄뒪?댁뒪??寃쎌슦
        elif hasattr(recv_comparison, "is_equal") and not isinstance(recv_comparison, dict):
            # QueryDifference ?대옒?ㅼ쓽 __str__ 硫붿꽌???ъ슜?섏뿬 "?쇱튂" ?먮뒗 "遺덉씪移? 媛?몄삤湲?            result_str = str(recv_comparison)
            cell = sheet.cell(row=row, column=10, value=result_str)
            # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
            if result_str == "?쇱튂":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
            elif result_str == "遺덉씪移?:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            else:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        # ?뺤뀛?덈━??寃쎌슦 ?뺤씤
        elif isinstance(recv_comparison, dict):
            if "is_equal" in recv_comparison:
                is_equal = recv_comparison.get("is_equal", False)
                result_str = "?쇱튂" if is_equal else "遺덉씪移?
                cell = sheet.cell(row=row, column=10, value=result_str)
                # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
                if is_equal:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            # ?먮뒗 "detail" ?꾨뱶媛 ?덉쓣 寃쎌슦
            elif "detail" in recv_comparison:
                detail = recv_comparison.get("detail", "")
                if detail == "?쇱튂" or detail == "遺덉씪移?:
                    cell = sheet.cell(row=row, column=10, value=detail)
                    # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
                    if detail == "?쇱튂":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??                else:
                    cell = sheet.cell(row=row, column=10, value="遺덉씪移?)
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??            else:
                cell = sheet.cell(row=row, column=10, value="鍮꾧탳遺덇?")
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        else:
            cell = sheet.cell(row=row, column=10, value="鍮꾧탳遺덇?")
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ?몃???        cell.font = font_normal

    def save_excel_output(self, output_path):
        """
        泥섎━??寃곌낵瑜??묒? ?뚯씪濡????        
        Args:
            output_path (str): 異쒕젰 ?묒? ?뚯씪 寃쎈줈
        
        Returns:
            bool: ????깃났 ?щ?
        """
        try:
            # ?뚯씪???대? ?대젮?덈뒗 寃쎌슦瑜??鍮꾪빐 ?덉쇅 泥섎━
            try:
                self.workbook.save(output_path)
                print(f"[寃곌낵 ????꾨즺] ?뚯씪 寃쎈줈: {output_path}")
                return True
            except PermissionError:
                print(f"[????ㅽ뙣] ?뚯씪???ㅻⅨ ?꾨줈洹몃옩?먯꽌 ?ъ슜 以묒엯?덈떎: {output_path}")
                return False
        except Exception as e:
            print(f"[????ㅽ뙣] ?ㅻ쪟 諛쒖깮: {str(e)}")
            return False

    def create_interface_sheet(self, if_info, mq_files=None, bw_files=None, queries=None, comparison_results=None):
        """
        ?묒? ?뚯씪??媛??명꽣?섏씠?ㅻ퀎 ?쒗듃瑜??앹꽦?섍퀬, ?곗씠?곕? 湲곕줉
        
        Args:
            if_info (dict): ?명꽣?섏씠???뺣낫
            mq_files (dict): MQ ?뚯씪 ?뺣낫 (?≪떊/?섏떊)
            bw_files (dict): BW ?뚯씪 ?뺣낫 (?≪떊/?섏떊)
            queries (dict): 荑쇰━ ?뺣낫 (MQ/BW, ?≪떊/?섏떊)
            comparison_results (dict): 鍮꾧탳 寃곌낵
        """
        if mq_files is None:
            mq_files = {}
        if bw_files is None:
            bw_files = {}
        if queries is None:
            queries = {}
        if comparison_results is None:
            comparison_results = {}
        
        # ?쒗듃 ?대쫫 ?앹꽦 (?명꽣?섏씠???대쫫 ?먮뒗 ID)
        sheet_name = if_info.get('interface_name', '') or if_info.get('interface_id', '')
        
        # ?쇰젴踰덊샇 李얘린 - ?붿빟 ?쒗듃?먯꽌 吏곸젒 李얘린
        seq_num = None
        interface_id = if_info.get('interface_id', '')
        
        if "?붿빟" in self.workbook.sheetnames:
            summary_sheet = self.workbook["?붿빟"]
            
            # ?붿빟 ?쒗듃?먯꽌 紐⑤뱺 ?됱쓣 ?쒗쉶?섎ŉ ?대떦 ?명꽣?섏씠??ID 李얘린
            for row_idx in range(2, summary_sheet.max_row + 1):
                cell_interface_id = summary_sheet.cell(row=row_idx, column=2).value
                
                if cell_interface_id == interface_id:
                    # ?쇰젴踰덊샇 李얠쓬
                    seq_num = summary_sheet.cell(row=row_idx, column=1).value
                    break
        
        # ?쒗듃 ?대쫫 ?욎뿉 ?쇰젴踰덊샇瑜?遺숈엫 (?덈뒗 寃쎌슦?먮쭔)
        if seq_num:
            sheet_name = f"{seq_num}_{sheet_name}"
        
        # ?쒗듃 ?대쫫??30?먮? 珥덇낵?섎㈃ ?먮Ⅴ湲?(Excel ?쒗듃 ?대쫫 ?쒗븳)
        if len(sheet_name) > 30:
            sheet_name = sheet_name[:27] + '...'
        
        # ?쒗듃 ?대쫫??以묐났?섎뒗 寃쎌슦 泥섎━
        base_name = sheet_name
        counter = 1
        while sheet_name in self.workbook.sheetnames:
            sheet_name = f"{base_name[:25]}_{counter}"
            counter += 1
        
        # ?명꽣?섏씠???쒗듃 ?앹꽦
        sheet = self.workbook.create_sheet(title=sheet_name)
        
        # ?ㅽ????뺤쓽
        header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        wrap_text_top = Alignment(wrap_text=True, vertical='top')
        
        # ???덈퉬 ?ㅼ젙 - 紐⑤뱺 ?댁쓣 A?닿낵 ?숈씪?섍쾶 ?ㅼ젙?섍퀬 2諛??ш린濡??뺤옣
        column_width = 30  # A?댁쓽 湲곕낯 ?덈퉬??2諛?        for col_letter in ['A', 'B', 'C', 'D', 'F']:
            sheet.column_dimensions[col_letter].width = column_width
        
        # E?댁? ?ㅻⅨ ?댁쓽 ?덈컲 ?ш린濡??ㅼ젙
        sheet.column_dimensions['E'].width = column_width / 2
        
        # 湲곕낯 湲瑗??ш린 ?ㅼ젙
        font_size_normal = 10
        font_size_query = 9
        
        # 1. ?명꽣?섏씠???뺣낫 ?ㅻ뜑 ?ㅼ젙
        row = 1
        sheet.cell(row=row, column=1, value="?명꽣?섏씠???뺣낫").font = Font(bold=True, size=font_size_normal)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        sheet.cell(row=row, column=1).fill = header_fill
        sheet.cell(row=row, column=1).alignment = align_center
        
        # ?명꽣?섏씠??ID 諛??대쫫
        row = 2
        sheet.cell(row=row, column=1, value="?명꽣?섏씠??ID").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=2, value=if_info.get('interface_id', ''))
        sheet.cell(row=row, column=3, value="?명꽣?섏씠??紐?).font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=4, value=if_info.get('interface_name', ''))
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        
        # ?≪떊 ?쒖뒪???뺣낫
        row = 3
        send_db_info = if_info.get('send', {}).get('db_info', {})
        send_sid = ''
        if isinstance(send_db_info, dict) and 'sid' in send_db_info:
            # sid?먯꽌 ip:port ?뺤떇 ?쒖쇅
            sid_full = send_db_info['sid']
            if isinstance(sid_full, str) and ':' in sid_full:
                # ip:port/sid ?뺤떇?대㈃ sid留?異붿텧
                send_sid = sid_full.split('/')[-1] if '/' in sid_full else sid_full
            else:
                send_sid = sid_full
                
        sheet.cell(row=row, column=1, value="?≪떊 ?쒖뒪??).font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=2, value=send_sid)
        sheet.cell(row=row, column=3, value="?≪떊 ?뚯씠釉?).font = Font(bold=True, size=font_size_normal)
        
        send_owner = if_info.get('send', {}).get('owner', '')
        send_table = if_info.get('send', {}).get('table_name', '')
        full_table_name = f"{send_owner}.{send_table}" if send_owner and send_table else send_table
        
        sheet.cell(row=row, column=4, value=full_table_name)
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        
        # ?섏떊 ?쒖뒪???뺣낫
        row = 4
        recv_db_info = if_info.get('recv', {}).get('db_info', {})
        recv_sid = ''
        if isinstance(recv_db_info, dict) and 'sid' in recv_db_info:
            # sid?먯꽌 ip:port ?뺤떇 ?쒖쇅
            sid_full = recv_db_info['sid']
            if isinstance(sid_full, str) and ':' in sid_full:
                # ip:port/sid ?뺤떇?대㈃ sid留?異붿텧
                recv_sid = sid_full.split('/')[-1] if '/' in sid_full else sid_full
            else:
                recv_sid = sid_full
                
        sheet.cell(row=row, column=1, value="?섏떊 ?쒖뒪??).font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=2, value=recv_sid)
        sheet.cell(row=row, column=3, value="?섏떊 ?뚯씠釉?).font = Font(bold=True, size=font_size_normal)
        
        recv_owner = if_info.get('recv', {}).get('owner', '')
        recv_table = if_info.get('recv', {}).get('table_name', '')
        full_recv_table = f"{recv_owner}.{recv_table}" if recv_owner and recv_table else recv_table
        
        sheet.cell(row=row, column=4, value=full_recv_table)
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        
        # 2. ?뚯씪 ?뺣낫 諛?荑쇰━ 鍮꾧탳 ?뱀뀡
        row = 6
        
        # 2.1 ?≪떊 ?뚯씪 ?뺣낫
        sheet.cell(row=row, column=1, value="?≪떊 ?뚯씪紐?).font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=1).fill = header_fill
        
        sheet.cell(row=row, column=2, value="MQ ?≪떊 ?뚯씪").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=2).fill = header_fill
        sheet.cell(row=row, column=2).alignment = align_center
        
        sheet.cell(row=row, column=4, value="BW ?≪떊 ?뚯씪").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=4).fill = header_fill
        sheet.cell(row=row, column=4).alignment = align_center
        
        # ?≪떊 ?뚯씪 ?뺣낫 ?낅젰
        row = 7
        sheet.cell(row=row, column=2, value=mq_files.get('send', {}).get('path', 'N/A'))
        if isinstance(bw_files, dict):
            sheet.cell(row=row, column=4, value=bw_files.get('send', 'N/A'))
        else:
            sheet.cell(row=row, column=4, value='N/A')
        
        # 2.2 ?섏떊 ?뚯씪 ?뺣낫
        row = 8
        sheet.cell(row=row, column=1, value="?섏떊 ?뚯씪紐?).font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=1).fill = header_fill
        
        sheet.cell(row=row, column=2, value="MQ ?섏떊 ?뚯씪").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=2).fill = header_fill
        sheet.cell(row=row, column=2).alignment = align_center
        
        sheet.cell(row=row, column=4, value="BW ?섏떊 ?뚯씪").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=4).fill = header_fill
        sheet.cell(row=row, column=4).alignment = align_center
        
        # ?섏떊 ?뚯씪 ?뺣낫 ?낅젰
        row = 9
        sheet.cell(row=row, column=2, value=mq_files.get('recv', {}).get('path', 'N/A'))
        if isinstance(bw_files, dict):
            sheet.cell(row=row, column=4, value=bw_files.get('recv', 'N/A'))
        else:
            sheet.cell(row=row, column=4, value='N/A')
        
        # 2.3 ?≪떊 荑쇰━ 諛?鍮꾧탳 寃곌낵
        row = 11
        sheet.cell(row=row, column=1, value="?≪떊 MQ 荑쇰━").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=1).fill = header_fill
        sheet.cell(row=row, column=1).alignment = align_center
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        
        sheet.cell(row=row, column=3, value="?≪떊 BW 荑쇰━").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=3).fill = header_fill
        sheet.cell(row=row, column=3).alignment = align_center
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        
        sheet.cell(row=row, column=5, value="鍮꾧탳").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=5).fill = header_fill
        sheet.cell(row=row, column=5).alignment = align_center
        
        sheet.cell(row=row, column=6, value="鍮꾧탳 ?곸꽭").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=6).fill = header_fill
        sheet.cell(row=row, column=6).alignment = align_center
        
        # ?≪떊 荑쇰━ ?뺣낫 ?낅젰
        row = 12
        sheet.cell(row=row, column=1, value=mq_files.get('send', {}).get('query', queries.get('mq_send', 'N/A')))
        sheet.cell(row=row, column=1).alignment = wrap_text_top
        sheet.cell(row=row, column=1).font = Font(size=font_size_query)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        
        sheet.cell(row=row, column=3, value=queries.get('bw_send', 'N/A'))
        sheet.cell(row=row, column=3).alignment = wrap_text_top
        sheet.cell(row=row, column=3).font = Font(size=font_size_query)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        
        # 鍮꾧탳 寃곌낵 ?낅젰
        is_send_equal = False
        send_detail = 'N/A'
        if comparison_results and 'send' in comparison_results:
            if isinstance(comparison_results['send'], dict):
                is_send_equal = comparison_results['send'].get('is_equal', False)
                send_detail = comparison_results['send'].get('detail', 'N/A')
            elif hasattr(comparison_results['send'], 'is_equal'):
                is_send_equal = comparison_results['send'].is_equal
                send_detail = getattr(comparison_results['send'], 'detail', 'N/A')
                
        result_cell = sheet.cell(row=row, column=5, value='?쇱튂' if is_send_equal else '遺덉씪移?)
        result_cell.alignment = align_center
        # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
        if is_send_equal:
            result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
        else:
            result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??        
        sheet.cell(row=row, column=6, value=send_detail)
        sheet.cell(row=row, column=6).alignment = wrap_text_top
        
        # 荑쇰━ ???믪씠 ?ㅼ젙
        sheet.row_dimensions[row].height = 150
        
        # 2.4 ?섏떊 荑쇰━ 諛?鍮꾧탳 寃곌낵
        row = 14
        sheet.cell(row=row, column=1, value="?섏떊 MQ 荑쇰━").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=1).fill = header_fill
        sheet.cell(row=row, column=1).alignment = align_center
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        
        sheet.cell(row=row, column=3, value="?섏떊 BW 荑쇰━").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=3).fill = header_fill
        sheet.cell(row=row, column=3).alignment = align_center
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        
        sheet.cell(row=row, column=5, value="鍮꾧탳").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=5).fill = header_fill
        sheet.cell(row=row, column=5).alignment = align_center
        
        sheet.cell(row=row, column=6, value="鍮꾧탳 ?곸꽭").font = Font(bold=True, size=font_size_normal)
        sheet.cell(row=row, column=6).fill = header_fill
        sheet.cell(row=row, column=6).alignment = align_center
        
        # ?섏떊 荑쇰━ ?뺣낫 ?낅젰
        row = 15
        sheet.cell(row=row, column=1, value=mq_files.get('recv', {}).get('query', queries.get('mq_recv', 'N/A')))
        sheet.cell(row=row, column=1).alignment = wrap_text_top
        sheet.cell(row=row, column=1).font = Font(size=font_size_query)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        
        sheet.cell(row=row, column=3, value=queries.get('bw_recv', 'N/A'))
        sheet.cell(row=row, column=3).alignment = wrap_text_top
        sheet.cell(row=row, column=3).font = Font(size=font_size_query)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        
        # 鍮꾧탳 寃곌낵 ?낅젰
        is_recv_equal = False
        recv_detail = 'N/A'
        if comparison_results and 'recv' in comparison_results:
            if isinstance(comparison_results['recv'], dict):
                is_recv_equal = comparison_results['recv'].get('is_equal', False)
                recv_detail = comparison_results['recv'].get('detail', 'N/A')
            elif hasattr(comparison_results['recv'], 'is_equal'):
                is_recv_equal = comparison_results['recv'].is_equal
                recv_detail = getattr(comparison_results['recv'], 'detail', 'N/A')
                
        result_cell = sheet.cell(row=row, column=5, value='?쇱튂' if is_recv_equal else '遺덉씪移?)
        result_cell.alignment = align_center
        # 寃곌낵???곕Ⅸ ? ?됱긽 ?곸슜
        if is_recv_equal:
            result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ?뱀깋
        else:
            result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 鍮④컙??        
        sheet.cell(row=row, column=6, value=recv_detail)
        sheet.cell(row=row, column=6).alignment = wrap_text_top
        
        # 荑쇰━ ???믪씠 ?ㅼ젙
        sheet.row_dimensions[row].height = 150
        
        # 紐⑤뱺 ????뚮몢由??곸슜
        for row_cells in sheet.iter_rows(min_row=1, max_row=15, min_col=1, max_col=6):
            for cell in row_cells:
                cell.border = border
        
        return sheet
    
    def close(self):
        """
        由ъ냼???뺣━
        """
        if self.workbook:
            if self.output_path:
                try:
                    self.workbook.save(self.output_path)
                except:
                    pass


# 媛꾨떒???ъ슜 ?덉떆
def main():
    # ?덉떆 肄붾뱶
    excel_manager = ExcelManager()
    sheet = excel_manager.initialize_excel_output()
    
    # 泥ル쾲吏??명꽣?섏씠??    # ?섑뵆 ?곗씠?곕줈 ?붿빟 ?쒗듃 ?낅뜲?댄듃
    sample_data = {
        "interface_info": {
            "interface_id": "IF001", 
            "interface_name": "?뚯뒪???명꽣?섏씠??, 
            "send": {
                "owner": "OWNER",
                "table_name": "TEST_TABLE",
                "db_info": {
                    "sid": "10.10.10.10:1521/DEVDB",
                    "system": "媛쒕컻?쒖뒪??
                }
            },
            "recv": {
                "owner": "OWNER",
                "table_name": "TEST_RCV_TABLE",
                "db_info": {
                    "sid": "10.10.10.20:1521/PRODDB",
                    "system": "?댁쁺?쒖뒪??
                }
            }
        }, 
        "file_results": {
            "send": {"path": "test.SND.xml", "query": "SELECT * FROM OWNER.TEST_TABLE"}, 
            "recv": {"path": "test.RCV.xml", "query": "SELECT * FROM OWNER.TEST_RCV_TABLE"}
        }, 
        "bw_files": ["bw_mapping.xml", "bw_mapping2.xml"], 
        "bw_queries": {
            "send": "INSERT INTO OWNER.TEST_TABLE VALUES (:1, :2, :3)",
            "recv": "INSERT INTO OWNER.TEST_RCV_TABLE VALUES (:1, :2, :3)"
        },
        "comparisons": {
            "send": {"is_equal": True, "detail": "?쇱튂"}, 
            "recv": {"is_equal": False, "detail": "遺덉씪移?}
        }
    }
    excel_manager.update_summary_sheet(sample_data)
    
    # 泥ル쾲吏??명꽣?섏씠???쒗듃 ?앹꽦
    if_info = {
        'interface_id': 'IF001',
        'interface_name': '?뚯뒪???명꽣?섏씠??,
        'send': {
            'owner': 'OWNER',
            'table_name': 'TEST_TABLE',
            'db_info': {
                'sid': '10.10.10.10:1521/DEVDB',
                'system': '媛쒕컻?쒖뒪??
            }
        },
        'recv': {
            'owner': 'OWNER',
            'table_name': 'TEST_RCV_TABLE',
            'db_info': {
                'sid': '10.10.10.20:1521/PRODDB',
                'system': '?댁쁺?쒖뒪??
            }
        }
    }
    
    mq_files = {
        'send': {'path': 'test.SND.xml', 'query': 'SELECT * FROM OWNER.TEST_TABLE WHERE 1=1 AND col1 = :value1 AND col2 = :value2'},
        'recv': {'path': 'test.RCV.xml', 'query': 'SELECT * FROM OWNER.TEST_RCV_TABLE WHERE status = \'Y\''}
    }
    
    bw_files = {
        'send': 'bw_mapping.xml',
        'recv': 'bw_mapping2.xml'
    }
    
    queries = {
        'mq_send': 'SELECT * FROM OWNER.TEST_TABLE WHERE 1=1 AND col1 = :value1 AND col2 = :value2',
        'bw_send': 'INSERT INTO OWNER.TEST_TABLE (col1, col2, col3) VALUES (:1, :2, :3)',
        'mq_recv': 'SELECT * FROM OWNER.TEST_RCV_TABLE WHERE status = \'Y\'',
        'bw_recv': 'INSERT INTO OWNER.TEST_RCV_TABLE (col1, col2, status) VALUES (:1, :2, \'Y\')'
    }
    
    comparison_results = {
        'send': {
            'is_equal': True, 
            'detail': '?쇱튂'
        },
        'recv': {
            'is_equal': False, 
            'detail': '遺덉씪移?
        }
    }
    
    # 泥ル쾲吏??명꽣?섏씠???쒗듃 ?앹꽦
    excel_manager.create_interface_sheet(if_info, mq_files, bw_files, queries, comparison_results)
    
    # ??踰덉㎏ ?섑뵆 ?명꽣?섏씠???곗씠??異붽?
    sample_data2 = {
        "interface_info": {
            "interface_id": "IF002", 
            "interface_name": "??踰덉㎏ ?뚯뒪???명꽣?섏씠??, 
            "send": {
                "owner": "OWNER2",
                "table_name": "TEST_TABLE2",
                "db_info": {
                    "sid": "10.10.10.10:1521/DEVDB",
                    "system": "媛쒕컻?쒖뒪??
                }
            },
            "recv": {
                "owner": "OWNER2",
                "table_name": "TEST_RCV_TABLE2",
                "db_info": {
                    "sid": "10.10.10.20:1521/PRODDB",
                    "system": "?댁쁺?쒖뒪??
                }
            }
        }, 
        "file_results": {
            "send": {"path": "test2.SND.xml", "query": "SELECT * FROM OWNER2.TEST_TABLE2"}, 
            "recv": {"path": "test2.RCV.xml", "query": "SELECT * FROM OWNER2.TEST_RCV_TABLE2"}
        }, 
        "bw_files": ["bw_mapping2.xml", "bw_mapping2_2.xml"], 
        "comparisons": {
            "send": {"is_equal": True, "detail": "?쇱튂"}, 
            "recv": {"is_equal": True, "detail": "?쇱튂"}
        }
    }
    excel_manager.update_summary_sheet(sample_data2)
    
    # ??踰덉㎏ ?명꽣?섏씠???쒗듃 ?앹꽦
    if_info2 = {
        'interface_id': 'IF002',
        'interface_name': '??踰덉㎏ ?뚯뒪???명꽣?섏씠??,
        'send': {
            'owner': 'OWNER2',
            'table_name': 'TEST_TABLE2',
            'db_info': {
                'sid': '10.10.10.10:1521/DEVDB',
                'system': '媛쒕컻?쒖뒪??
            }
        },
        'recv': {
            'owner': 'OWNER2',
            'table_name': 'TEST_RCV_TABLE2',
            'db_info': {
                'sid': '10.10.10.20:1521/PRODDB',
                'system': '?댁쁺?쒖뒪??
            }
        }
    }
    
    mq_files2 = {
        'send': {'path': 'test2.SND.xml', 'query': 'SELECT * FROM OWNER2.TEST_TABLE2 WHERE 1=1'},
        'recv': {'path': 'test2.RCV.xml', 'query': 'SELECT * FROM OWNER2.TEST_RCV_TABLE2 WHERE status = \'Y\''}
    }
    
    # ??踰덉㎏ ?명꽣?섏씠???쒗듃 ?앹꽦
    excel_manager.create_interface_sheet(if_info2, mq_files2, bw_files, queries, comparison_results)
    
    # 寃곌낵 ???    output_path = 'comp_mq_bw_sample.xlsx'
    excel_manager.save_excel_output(output_path)
    print(f"寃곌낵媛 {output_path}????λ릺?덉뒿?덈떎.")
    
    excel_manager.close()


if __name__ == "__main__":
    main()
